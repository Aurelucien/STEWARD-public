"""Bounded, isolated Structured Document Observation foundations.

This module is the repository-owned filesystem-observation boundary selected
by the Structured Document Parser Contract.  It intentionally sits below the
model-visible Tool Registry: I1 establishes the offline PDF vertical slice,
not a new provider-facing tool.
"""

from __future__ import annotations

from collections.abc import Callable, Iterator
from contextlib import contextmanager, redirect_stderr, redirect_stdout
from dataclasses import dataclass, replace
import errno
import gc
from io import BytesIO
from hashlib import sha256
from importlib import import_module
from importlib.metadata import version
from html.parser import HTMLParser
import multiprocessing
import os
import posixpath
from pathlib import Path, PurePosixPath
import psutil
import queue
import re
import resource
import stat
import sys
from tempfile import TemporaryDirectory
from time import monotonic
from typing import Any, TypeAlias, cast
from urllib.parse import unquote
import zipfile
from xml.parsers import expat
from xml.etree import ElementTree
import math

from ...evidence import canonical_json
from ...document_execution import (
    BoundedDocumentParseCache,
    DocumentContainerQuality,
    DocumentExecutionAttempt,
    DocumentExecutionSelection,
    DocumentExecutionTrace,
    assess_document_quality,
    initial_document_profile,
)
from ...document_query import document_match_mode, match_document_text
from .docling_documents import (
    DoclingOcrPageRangeWorker,
    DoclingPageRangeWorker,
    docling_document_worker,
    docling_enriched_document_worker,
    docling_macos_ocr_worker,
)
from .failures import RuntimeFailure
from .native_fidelity import (
    MAX_OPTIONAL_XML_BYTES,
    chart_search_text,
    openpyxl_chart_extension,
    pptx_chart_extension,
    pptx_shape_accessibility,
    project_docx_auxiliary,
    project_pdf_native,
)
from .scope_binding import ScopeBindings
from .audio_alignment import (
    AudioAlignmentUnavailable,
    alignment_runtime_capabilities,
    resolve_local_alignment_model,
)
from .audio_diarization import (
    AudioDiarizationUnavailable,
    resolve_local_diarization_model,
)
from .audio_documents import (
    AUDIO_FORMAT_BY_SUFFIX,
    AUDIO_SOURCE_FORMATS,
    AUDIO_SUFFIX_BY_FORMAT,
    MAX_AUDIO_SOURCE_BYTES,
    AudioDocumentWorker,
    AudioRuntimeUnavailable,
    audio_request_digest,
    audio_runtime_capabilities,
    resolve_local_audio_model,
)
from .video_documents import (
    MAX_VIDEO_SOURCE_BYTES,
    VIDEO_RESULT_PAGE_CONTINUATION_SCHEMA_VERSION,
    VIDEO_FORMAT_BY_SUFFIX,
    VIDEO_SOURCE_FORMATS,
    VIDEO_SUFFIX_BY_FORMAT,
    VideoProbeWorker,
    VideoSceneWorker,
    VideoTimelineWorker,
    video_request_digest,
    video_runtime_capabilities,
)


# The absolute staging ceiling is deliberately separate from format admission.
# Sources are copied and hashed incrementally, so this is a disk/work bound rather
# than a Python heap allocation request.
MAX_SOURCE_BYTES = 1024 * 1024 * 1024
MAX_EXPANDED_BYTES = 4 * 1024 * 1024 * 1024
MAX_PDF_SOURCE_BYTES = 1024 * 1024 * 1024
MAX_PACKAGE_SOURCE_BYTES = 512 * 1024 * 1024
MAX_IMAGE_SOURCE_BYTES = 512 * 1024 * 1024
MAX_PACKAGE_EXPANDED_BYTES = 2 * 1024 * 1024 * 1024
MAX_XLSX_EXPANDED_BYTES = MAX_EXPANDED_BYTES
MAX_PACKAGE_MEMBERS = 100_000
MAX_PACKAGE_MEMBER_BYTES = 1024 * 1024 * 1024
MAX_PACKAGE_COMPRESSION_RATIO = 1_000
MAX_CONTROL_XML_BYTES = 8 * 1024 * 1024
MAX_RELATIONSHIP_PARTS = 10_000
MAX_RELATIONSHIP_BYTES = 64 * 1024 * 1024
DOCUMENT_INGRESS_CHUNK_BYTES = 1024 * 1024
STREAMING_QUERY_MAP_THRESHOLD_BYTES = 32 * 1024 * 1024
STREAMING_QUERY_MAP_PDF_PAGE_THRESHOLD = 128
MAX_STREAMING_QUERY_MATCH_ITEMS = 512
MAX_STREAMING_QUERY_EXCERPT_CHARS = 2_048
PDF_PAGE_OCR_RENDER_SCALE = 1.0
MAX_PDF_PAGE_OCR_PIXELS = 4_000_000
MIN_PDF_NATIVE_TEXT_CHARACTERS = 8
MAX_EPUB_NATIVE_READ_TEXT_BYTES = 512 * 1024
MAX_EPUB_NATIVE_ITEMS = 3_000
MAX_NATIVE_FORMULA_ITEMS = 1_000
MAX_ADAPTIVE_PARSER_ELAPSED_SECONDS = 600.0
DOCUMENT_OPERATION_RELEASE_GRACE_SECONDS = 60.0
MAX_DOCUMENT_OPERATION_ELAPSED_SECONDS = (
    MAX_ADAPTIVE_PARSER_ELAPSED_SECONDS + DOCUMENT_OPERATION_RELEASE_GRACE_SECONDS
)
MAX_PARSER_ELAPSED_SECONDS = 6.0
MAX_PARSER_MEMORY_BYTES = 640 * 1024 * 1024
MAX_PDF_PARSER_ELAPSED_SECONDS = 45.0
MAX_PDF_PARSER_MEMORY_BYTES = 2 * 1024 * 1024 * 1024
MAX_DEEP_PARSER_ELAPSED_SECONDS = 120.0
MAX_DEEP_PARSER_MEMORY_BYTES = 3 * 1024 * 1024 * 1024
MAX_PARSED_ITEMS_OR_BLOCKS = 20_000
MAX_NORMALIZED_OUTPUT_BYTES = 2 * 1024 * 1024
IMAGE_SOURCE_FORMATS = frozenset({"PNG", "JPEG", "TIFF"})
MAX_TARGETED_EVIDENCE_PAGES = 8

CURRENT_FILESYSTEM_DOCUMENT = "CURRENT_FILESYSTEM_DOCUMENT"
CURRENT_FILESYSTEM_AUDIO = "CURRENT_FILESYSTEM_AUDIO"
CURRENT_FILESYSTEM_VIDEO = "CURRENT_FILESYSTEM_VIDEO"

WorkerTarget: TypeAlias = Callable[[str], dict[str, Any]]

_DOCX_MAIN_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"
)
_DOCX_MACRO_CONTENT_TYPE = "application/vnd.ms-word.document.macroEnabled.main+xml"
_DOCX_DATA_URL_PATTERN = re.compile(
    r"data:[A-Za-z0-9.+-]+/[A-Za-z0-9.+-]+(?:;[A-Za-z0-9.+-]+)*(?:,[^\s)]*|\.\.\.)?",
    re.IGNORECASE,
)
_DOCX_BINARY_LINE_PATTERN = re.compile(r"[A-Za-z0-9+/]{128,}={0,2}")
_DOCX_HEADING_PATTERN = re.compile(r"^(#{1,6})\s+(.+)$")


class _RecognizedDocumentMalformedError(Exception):
    """A recognized document container has invalid backend-required structure."""


@dataclass(frozen=True, slots=True)
class DocumentResourceUsage:
    """Measured document-adapter resource facts; all byte values are explicit."""

    source_bytes: int | None
    expanded_bytes: int
    parser_elapsed_ms: int
    parser_memory_bytes: int | None
    parsed_items_or_blocks: int
    normalized_output_bytes: int
    admission_profile: str | None = None
    source_limit_bytes: int | None = None
    expanded_limit_bytes: int | None = None
    archive_member_count: int | None = None
    parser_timeout_limit_ms: int | None = None
    parser_memory_limit_bytes: int | None = None
    ingress_elapsed_ms: int = 0
    identification_elapsed_ms: int = 0
    operation_elapsed_ms: int = 0
    deadline_stage: str | None = None
    media: dict[str, object] | None = None

    def payload(self) -> dict[str, object]:
        value: dict[str, object] = {
            "source_bytes": self.source_bytes,
            "expanded_bytes": self.expanded_bytes,
            "parser_elapsed_ms": self.parser_elapsed_ms,
            "parser_memory_bytes": self.parser_memory_bytes,
            "parsed_items_or_blocks": self.parsed_items_or_blocks,
            "normalized_output_bytes": self.normalized_output_bytes,
        }
        if self.admission_profile is not None:
            value["admission_profile"] = self.admission_profile
        if self.source_limit_bytes is not None:
            value["source_limit_bytes"] = self.source_limit_bytes
        if self.expanded_limit_bytes is not None:
            value["expanded_limit_bytes"] = self.expanded_limit_bytes
        if self.archive_member_count is not None:
            value["archive_member_count"] = self.archive_member_count
        if self.parser_timeout_limit_ms is not None:
            value["parser_timeout_limit_ms"] = self.parser_timeout_limit_ms
        if self.parser_memory_limit_bytes is not None:
            value["parser_memory_limit_bytes"] = self.parser_memory_limit_bytes
        value["ingress_elapsed_ms"] = self.ingress_elapsed_ms
        value["identification_elapsed_ms"] = self.identification_elapsed_ms
        value["operation_elapsed_ms"] = self.operation_elapsed_ms
        if self.deadline_stage is not None:
            value["deadline_stage"] = self.deadline_stage
        if self.media is not None:
            value["media"] = self.media
        return value


@dataclass(frozen=True, slots=True)
class DocumentSourceProvenance:
    """Safe current-filesystem provenance without a host path."""

    scope_id: str
    relative_path: str
    source_sha256: str | None
    source_kind: str = CURRENT_FILESYSTEM_DOCUMENT

    def payload(self) -> dict[str, str | None]:
        return {
            "source_kind": self.source_kind,
            "scope_id": self.scope_id,
            "relative_path": self.relative_path,
            "source_sha256": self.source_sha256,
        }


@dataclass(frozen=True, slots=True)
class NormalizedDocumentItem:
    """One backend-neutral document-relative logical item."""

    kind: str
    text_or_value: str | None
    parent: str | None
    location: dict[str, int | str]
    extension: dict[str, object] | None = None
    node_id: str | None = None
    role: str | None = None

    def payload(self) -> dict[str, object]:
        value: dict[str, object] = {"kind": self.kind, "location": self.location}
        if self.node_id is not None:
            value["node_id"] = self.node_id
        if self.role is not None:
            value["role"] = self.role
        if self.text_or_value is not None:
            value["text_or_value"] = self.text_or_value
        if self.parent is not None:
            value["parent"] = self.parent
        if self.extension is not None:
            value["extension"] = self.extension
        return value


@dataclass(frozen=True, slots=True)
class NormalizedDocumentObservation:
    """Provider-neutral normalized current document observation."""

    status: str
    source_format: str | None
    backend_name: str | None
    backend_version: str | None
    provenance: DocumentSourceProvenance
    warnings: tuple[str, ...]
    items: tuple[NormalizedDocumentItem, ...]
    resources: DocumentResourceUsage
    identification_reason: str | None = None
    execution: DocumentExecutionTrace | None = None
    continuation: dict[str, object] | None = None
    failure_reason_code: str | None = None
    failure_exception_type: str | None = None

    def payload(self) -> dict[str, object]:
        value: dict[str, object] = {
            "status": self.status,
            "source_format": self.source_format,
            "backend_name": self.backend_name,
            "backend_version": self.backend_version,
            "source_provenance": self.provenance.payload(),
            "warnings": list(self.warnings),
            "items": [item.payload() for item in self.items],
            "resource_usage": self.resources.payload(),
        }
        if self.identification_reason is not None:
            value["identification_reason"] = self.identification_reason
        if self.execution is not None:
            value["execution"] = self.execution.payload()
        if self.continuation is not None:
            value["continuation"] = self.continuation
        if self.failure_reason_code is not None:
            value["failure_reason_code"] = self.failure_reason_code
        if self.failure_exception_type is not None:
            value["failure_exception_type"] = self.failure_exception_type
        return value

    @property
    def result_digest(self) -> str:
        return sha256(canonical_json(self.payload())).hexdigest()


@dataclass(frozen=True, slots=True)
class _IngressFailure:
    status: str
    scope_id: str
    relative_path: str
    source_bytes: int | None


@dataclass(slots=True)
class _AdmittedDocumentSource:
    """One incrementally staged source; no user path crosses this boundary."""

    scope_id: str
    relative_path: str
    source_bytes: int
    source_sha256: str
    _temporary: TemporaryDirectory[str]
    _staged_path: Path
    admission_profile: str | None = None
    source_limit_bytes: int | None = None
    expanded_limit_bytes: int | None = None
    archive_member_count: int | None = None
    ingress_elapsed_ms: int = 0
    identification_elapsed_ms: int = 0

    def close(self) -> None:
        self._temporary.cleanup()

    @contextmanager
    def staged_copy(self, suffix: str) -> Iterator[Path]:
        """Offer a suffix-correct hard link without duplicating staged content."""
        path = self._staged_path.with_name(f"source{suffix}")
        if not path.exists():
            try:
                os.link(self._staged_path, path)
            except OSError:
                with self._staged_path.open("rb") as source, path.open("xb") as target:
                    while chunk := source.read(DOCUMENT_INGRESS_CHUNK_BYTES):
                        target.write(chunk)
            path.chmod(stat.S_IRUSR)
        yield path


def _same_state(before: os.stat_result, after: os.stat_result) -> bool:
    return (
        before.st_dev,
        before.st_ino,
        before.st_size,
        before.st_mtime_ns,
    ) == (
        after.st_dev,
        after.st_ino,
        after.st_size,
        after.st_mtime_ns,
    )


@dataclass(slots=True)
class ProjectOwnedBoundedDocumentIngress:
    """Descriptor-safe document ingress; it is not an AgentRuntime filesystem bypass."""

    bindings: ScopeBindings
    read_bytes: Callable[[int, int], bytes] = os.read
    require_same_device: bool = False
    max_staged_bytes: int = MAX_SOURCE_BYTES

    def preflight(self, arguments: dict[str, object]) -> None:
        scope_id, relative_path = _scope_arguments(arguments)
        self.bindings.require(scope_id).resolve_relative_path(relative_path)

    def admit(self, arguments: dict[str, object]) -> _AdmittedDocumentSource | _IngressFailure:
        scope_id, relative_path = _scope_arguments(arguments)
        binding = self.bindings.require(scope_id)
        binding.resolve_relative_path(relative_path)
        try:
            return self._read(binding.allowed_root, scope_id, relative_path)
        except OSError:
            return _IngressFailure("UNAVAILABLE", scope_id, relative_path, None)

    def _read(
        self, root: Path, scope_id: str, relative_path: str
    ) -> _AdmittedDocumentSource | _IngressFailure:
        started = monotonic()
        root_resolved = root.resolve(strict=True)
        root_flags = os.O_RDONLY | getattr(os, "O_DIRECTORY", 0)
        nofollow = getattr(os, "O_NOFOLLOW", None)
        if nofollow is None:
            raise OSError(errno.ENOTSUP, "descriptor-safe no-follow open is unavailable")
        root_fd = os.open(root_resolved, root_flags)
        opened = [root_fd]
        try:
            root_state = os.fstat(root_fd)
            parent_fd = root_fd
            components = relative_path.split("/")
            for component in components[:-1]:
                child_fd = os.open(component, root_flags | nofollow, dir_fd=parent_fd)
                opened.append(child_fd)
                parent_fd = child_fd
            source_fd = os.open(components[-1], os.O_RDONLY | nofollow, dir_fd=parent_fd)
            opened.append(source_fd)
            before = os.fstat(source_fd)
            if not stat.S_ISREG(before.st_mode):
                return _IngressFailure("UNAVAILABLE", scope_id, relative_path, None)
            if self.require_same_device and before.st_dev != root_state.st_dev:
                return _IngressFailure("UNAVAILABLE", scope_id, relative_path, None)
            if before.st_size > self.max_staged_bytes:
                return _IngressFailure("RESOURCE_LIMIT", scope_id, relative_path, before.st_size)
            temporary = TemporaryDirectory(prefix="steward-document-")
            staged_path = Path(temporary.name) / "source.bin"
            digest = sha256()
            observed_bytes = 0
            try:
                with staged_path.open("xb") as staged:
                    while True:
                        chunk = self.read_bytes(source_fd, DOCUMENT_INGRESS_CHUNK_BYTES)
                        if not chunk:
                            break
                        observed_bytes += len(chunk)
                        if observed_bytes > self.max_staged_bytes:
                            temporary.cleanup()
                            return _IngressFailure(
                                "RESOURCE_LIMIT", scope_id, relative_path, observed_bytes
                            )
                        digest.update(chunk)
                        staged.write(chunk)
                staged_path.chmod(stat.S_IRUSR)
            except Exception:
                temporary.cleanup()
                raise
            after = os.fstat(source_fd)
            if not _same_state(before, after):
                temporary.cleanup()
                return _IngressFailure("UNAVAILABLE", scope_id, relative_path, None)
            return _AdmittedDocumentSource(
                scope_id,
                relative_path,
                observed_bytes,
                digest.hexdigest(),
                temporary,
                staged_path,
                ingress_elapsed_ms=int((monotonic() - started) * 1_000),
            )
        finally:
            for descriptor in reversed(opened):
                os.close(descriptor)


def _scope_arguments(arguments: dict[str, object]) -> tuple[str, str]:
    scope_id = arguments.get("scope_id")
    relative_path = arguments.get("relative_path")
    if not isinstance(scope_id, str) or not isinstance(relative_path, str):
        raise RuntimeFailure("SCOPE_BINDING_FAILED", "scope and relative path are required")
    return scope_id, relative_path


@dataclass(frozen=True, slots=True)
class _FormatIdentification:
    source_format: str | None
    status: str | None
    reason: str | None
    expanded_bytes: int = 0
    external_relationship_count: int = 0
    archive_member_count: int = 0
    admission_profile: str | None = None
    source_limit_bytes: int | None = None
    expanded_limit_bytes: int | None = None


DocumentFormatSource: TypeAlias = bytes | Path


def _source_prefix(source: DocumentFormatSource, length: int = 16) -> bytes:
    if isinstance(source, bytes):
        return source[:length]
    with source.open("rb") as handle:
        return handle.read(length)


def _source_size(source: DocumentFormatSource) -> int:
    return len(source) if isinstance(source, bytes) else source.stat().st_size


def _format_budget(source_format: str) -> tuple[str, int, int | None]:
    if source_format in AUDIO_SOURCE_FORMATS:
        return "AUDIO_STREAM", MAX_AUDIO_SOURCE_BYTES, None
    if source_format in VIDEO_SOURCE_FORMATS:
        return "VIDEO_STREAM", MAX_VIDEO_SOURCE_BYTES, None
    if source_format == "PDF":
        return "PAGINATED_STREAM", MAX_PDF_SOURCE_BYTES, None
    if source_format in IMAGE_SOURCE_FORMATS:
        return "RASTER_STREAM", MAX_IMAGE_SOURCE_BYTES, None
    if source_format == "XLSX":
        return "PACKAGE_STREAM", MAX_PACKAGE_SOURCE_BYTES, MAX_XLSX_EXPANDED_BYTES
    return "PACKAGE_STREAM", MAX_PACKAGE_SOURCE_BYTES, MAX_PACKAGE_EXPANDED_BYTES


def _identified_format(
    source_format: str,
    *,
    expanded_bytes: int = 0,
    external_relationship_count: int = 0,
    archive_member_count: int = 0,
) -> _FormatIdentification:
    profile, source_limit, expanded_limit = _format_budget(source_format)
    return _FormatIdentification(
        source_format,
        None,
        None,
        expanded_bytes,
        external_relationship_count,
        archive_member_count,
        profile,
        source_limit,
        expanded_limit,
    )


def identify_document_format(
    source: DocumentFormatSource, relative_path: str
) -> _FormatIdentification:
    """Apply bounded signature evidence; filename suffixes never route a backend alone."""
    prefix = _source_prefix(source)
    source_bytes = _source_size(source)
    for signature, source_format in (
        (b"%PDF-", "PDF"),
        (b"\x89PNG\r\n\x1a\n", "PNG"),
        (b"\xff\xd8\xff", "JPEG"),
        (b"II*\x00", "TIFF"),
        (b"MM\x00*", "TIFF"),
    ):
        if prefix.startswith(signature):
            identified = _identified_format(source_format)
            if (
                identified.source_limit_bytes is not None
                and source_bytes > identified.source_limit_bytes
            ):
                return replace(
                    identified,
                    status="RESOURCE_LIMIT",
                    reason="SOURCE_BYTES_LIMIT",
                )
            return identified
    archive_source: bytes | Path = source
    zip_candidate = BytesIO(source) if isinstance(source, bytes) else source
    if zipfile.is_zipfile(zip_candidate):
        return _identify_zip_container(archive_source)
    suffix = Path(relative_path).suffix.lower()
    if suffix in AUDIO_FORMAT_BY_SUFFIX:
        identified = _identified_format(AUDIO_FORMAT_BY_SUFFIX[suffix])
        if source_bytes > MAX_AUDIO_SOURCE_BYTES:
            return replace(identified, status="RESOURCE_LIMIT", reason="SOURCE_BYTES_LIMIT")
        # The suffix only admits the file to the bounded ffprobe boundary; the
        # parser does not publish an audio identity until ffprobe verifies it.
        return identified
    if suffix in VIDEO_FORMAT_BY_SUFFIX:
        identified = _identified_format(VIDEO_FORMAT_BY_SUFFIX[suffix])
        if source_bytes > MAX_VIDEO_SOURCE_BYTES:
            return replace(identified, status="RESOURCE_LIMIT", reason="SOURCE_BYTES_LIMIT")
        # The suffix only admits the file to FFprobe. The demuxer and a usable
        # non-attached video stream establish the actual media identity.
        return identified
    if suffix == ".pdf":
        return _FormatIdentification(None, "UNSUPPORTED_FORMAT", "FORMAT_MISMATCH")
    if Path(relative_path).suffix.lower() in {
        ".epub",
        ".docx",
        ".xlsx",
        ".pptx",
        ".png",
        ".jpg",
        ".jpeg",
        ".tif",
        ".tiff",
        *AUDIO_FORMAT_BY_SUFFIX,
        *VIDEO_FORMAT_BY_SUFFIX,
    }:
        return _FormatIdentification(None, "UNSUPPORTED_FORMAT", "FORMAT_MISMATCH")
    return _FormatIdentification(None, "UNSUPPORTED_FORMAT", "UNKNOWN_INPUT")


def _archive_input(source: DocumentFormatSource) -> BytesIO | Path:
    return BytesIO(source) if isinstance(source, bytes) else source


def _bounded_archive_read(
    archive: zipfile.ZipFile, member: str | zipfile.ZipInfo, limit: int = MAX_CONTROL_XML_BYTES
) -> bytes:
    with archive.open(member) as handle:
        value = handle.read(limit + 1)
    if len(value) > limit:
        raise _RecognizedDocumentMalformedError("container control part exceeds its bound")
    return value


def _container_safety_failure(
    source_format: str, members: list[zipfile.ZipInfo], expanded_limit: int
) -> _FormatIdentification | None:
    if len(members) > MAX_PACKAGE_MEMBERS:
        return _FormatIdentification(
            source_format,
            "RESOURCE_LIMIT",
            "ARCHIVE_MEMBER_COUNT_LIMIT",
            archive_member_count=len(members),
        )
    names = [member.filename for member in members]
    if len(names) != len(set(names)):
        return _FormatIdentification(source_format, "MALFORMED", "DUPLICATE_CONTAINER_MEMBER")
    if any(member.flag_bits & 0x1 for member in members):
        return _FormatIdentification(source_format, "UNSUPPORTED_FORMAT", "ENCRYPTED_CONTAINER")
    if any(
        member.filename.startswith("/") or ".." in PurePosixPath(member.filename).parts
        for member in members
    ):
        return _FormatIdentification(source_format, "RESOURCE_LIMIT", "UNSAFE_CONTAINER_PATH")
    expanded_bytes = sum(member.file_size for member in members)
    if expanded_bytes > expanded_limit:
        return _FormatIdentification(
            source_format,
            "RESOURCE_LIMIT",
            "EXPANDED_BYTES_LIMIT",
            expanded_bytes,
            archive_member_count=len(members),
        )
    if any(member.file_size > MAX_PACKAGE_MEMBER_BYTES for member in members):
        return _FormatIdentification(
            source_format,
            "RESOURCE_LIMIT",
            "ARCHIVE_MEMBER_BYTES_LIMIT",
            expanded_bytes,
            archive_member_count=len(members),
        )
    relationship_parts = [member for member in members if member.filename.endswith(".rels")]
    if (
        len(relationship_parts) > MAX_RELATIONSHIP_PARTS
        or sum(member.file_size for member in relationship_parts) > MAX_RELATIONSHIP_BYTES
    ):
        return _FormatIdentification(
            source_format,
            "RESOURCE_LIMIT",
            "RELATIONSHIP_PARTS_LIMIT",
            expanded_bytes,
            archive_member_count=len(members),
        )
    control_names = {"[Content_Types].xml", "META-INF/container.xml"}
    if any(
        member.filename in control_names and member.file_size > MAX_CONTROL_XML_BYTES
        for member in members
    ):
        return _FormatIdentification(
            source_format,
            "RESOURCE_LIMIT",
            "CONTROL_PART_BYTES_LIMIT",
            expanded_bytes,
            archive_member_count=len(members),
        )
    for member in members:
        if member.file_size <= MAX_CONTROL_XML_BYTES:
            continue
        if member.compress_size == 0 or member.file_size > (
            member.compress_size * MAX_PACKAGE_COMPRESSION_RATIO
        ):
            return _FormatIdentification(
                source_format,
                "RESOURCE_LIMIT",
                "ARCHIVE_COMPRESSION_RATIO_LIMIT",
                expanded_bytes,
                archive_member_count=len(members),
            )
    return None


def _identify_zip_container(source: DocumentFormatSource) -> _FormatIdentification:
    """Inspect bounded EPUB/OOXML evidence before routing a container backend."""
    try:
        with zipfile.ZipFile(_archive_input(source)) as archive:
            members = archive.infolist()
            names = {member.filename for member in members}
            if "mimetype" in names:
                try:
                    media_type = _bounded_archive_read(archive, "mimetype", 64)
                except KeyError:
                    media_type = b""
                if media_type == b"application/epub+zip":
                    return _identify_epub_container(
                        archive, members, names, source_bytes=_source_size(source)
                    )
            formats = {
                "DOCX": "word/document.xml" in names,
                "XLSX": "xl/workbook.xml" in names,
                "PPTX": "ppt/presentation.xml" in names,
            }
            recognized = [format_name for format_name, present in formats.items() if present]
            if not recognized:
                return _FormatIdentification(None, "UNSUPPORTED_FORMAT", "UNACCEPTED_FORMAT")
            if len(recognized) != 1:
                return _FormatIdentification(None, "UNSUPPORTED_FORMAT", "FORMAT_MISMATCH")
            source_format = recognized[0]
            profile, source_limit, expanded_limit = _format_budget(source_format)
            if _source_size(source) > source_limit:
                return _FormatIdentification(
                    source_format,
                    "RESOURCE_LIMIT",
                    "SOURCE_BYTES_LIMIT",
                    archive_member_count=len(members),
                    admission_profile=profile,
                    source_limit_bytes=source_limit,
                    expanded_limit_bytes=expanded_limit,
                )
            if "[Content_Types].xml" not in names:
                return _FormatIdentification(source_format, "MALFORMED", "MISSING_CONTENT_TYPES")
            safety_failure = _container_safety_failure(
                source_format, members, expanded_limit or MAX_EXPANDED_BYTES
            )
            if safety_failure is not None:
                return replace(
                    safety_failure,
                    admission_profile=profile,
                    source_limit_bytes=source_limit,
                    expanded_limit_bytes=expanded_limit,
                )
            expanded_bytes = sum(member.file_size for member in members)
            if source_format == "DOCX":
                docx_identity = _docx_package_identity(archive)
                if docx_identity == "MACRO_ENABLED":
                    return _FormatIdentification(None, "UNSUPPORTED_FORMAT", "UNACCEPTED_FORMAT")
                if docx_identity != "VALID":
                    return _FormatIdentification(
                        "DOCX", "MALFORMED", "INVALID_DOCX_PACKAGE_IDENTITY"
                    )
            external_relationship_count = _external_relationship_count(archive, members)
            return _identified_format(
                source_format,
                expanded_bytes=expanded_bytes,
                external_relationship_count=external_relationship_count,
                archive_member_count=len(members),
            )
    except (
        OSError,
        zipfile.BadZipFile,
        ElementTree.ParseError,
        _RecognizedDocumentMalformedError,
    ):
        return _FormatIdentification(None, "MALFORMED", "MALFORMED_CONTAINER")


def _identify_epub_container(
    archive: zipfile.ZipFile,
    members: list[zipfile.ZipInfo],
    names: set[str],
    *,
    source_bytes: int,
) -> _FormatIdentification:
    """Validate the safe structural minimum for an EPUB 2/3 package."""
    if "META-INF/container.xml" not in names:
        return _FormatIdentification("EPUB", "MALFORMED", "MISSING_EPUB_CONTAINER")
    if "META-INF/encryption.xml" in names:
        return _FormatIdentification("EPUB", "UNSUPPORTED_FORMAT", "ENCRYPTED_EPUB")
    profile, source_limit, expanded_limit = _format_budget("EPUB")
    if source_bytes > source_limit:
        return _FormatIdentification(
            "EPUB",
            "RESOURCE_LIMIT",
            "SOURCE_BYTES_LIMIT",
            archive_member_count=len(members),
            admission_profile=profile,
            source_limit_bytes=source_limit,
            expanded_limit_bytes=expanded_limit,
        )
    safety_failure = _container_safety_failure("EPUB", members, expanded_limit or 0)
    if safety_failure is not None:
        return replace(
            safety_failure,
            admission_profile=profile,
            source_limit_bytes=source_limit,
            expanded_limit_bytes=expanded_limit,
        )
    expanded_bytes = sum(member.file_size for member in members)
    try:
        root = ElementTree.fromstring(_bounded_archive_read(archive, "META-INF/container.xml"))
    except (KeyError, ElementTree.ParseError):
        return _FormatIdentification("EPUB", "MALFORMED", "INVALID_EPUB_CONTAINER")
    rootfiles = [
        element.get("full-path")
        for element in root.iter()
        if element.tag.rsplit("}", 1)[-1] == "rootfile"
    ]
    package_paths = [path for path in rootfiles if isinstance(path, str) and path in names]
    if not package_paths:
        return _FormatIdentification("EPUB", "MALFORMED", "MISSING_EPUB_PACKAGE")
    return _identified_format(
        "EPUB", expanded_bytes=expanded_bytes, archive_member_count=len(members)
    )


def _docx_package_identity(archive: zipfile.ZipFile) -> str:
    """Distinguish DOCX from macro-enabled or structurally invalid Word packages."""
    try:
        root = ElementTree.fromstring(_bounded_archive_read(archive, "[Content_Types].xml"))
    except (KeyError, ElementTree.ParseError):
        return "INVALID"
    content_types = {element.get("ContentType") for element in root}
    if _DOCX_MAIN_CONTENT_TYPE in content_types:
        return "VALID"
    if _DOCX_MACRO_CONTENT_TYPE in content_types:
        return "MACRO_ENABLED"
    return "INVALID"


def _external_relationship_count(archive: zipfile.ZipFile, members: list[zipfile.ZipInfo]) -> int:
    """Count, but never follow, relationships declared external by OOXML."""
    count = 0
    for member in members:
        if not member.filename.endswith(".rels"):
            continue
        try:
            root = ElementTree.fromstring(_bounded_archive_read(archive, member))
        except ElementTree.ParseError:
            raise
        count += sum(relation.get("TargetMode") == "External" for relation in root)
    return count


@dataclass(frozen=True, slots=True)
class _WorkerExecution:
    status: str
    payload: dict[str, Any] | None
    elapsed_ms: int
    peak_memory_bytes: int | None
    failure_reason_code: str | None = None
    failure_exception_type: str | None = None


def _safe_exception_type(error: Exception) -> str:
    """Return a bounded class name without messages, paths, or traceback content."""
    name = type(error).__name__
    return name[:128] if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", name) else "Exception"


def _worker_peak_memory_bytes() -> int | None:
    peak = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
    if peak <= 0:
        return None
    return peak if sys.platform == "darwin" else peak * 1024


def _worker_entry(
    target: WorkerTarget,
    source_path: str,
    result_queue: multiprocessing.queues.Queue[Any],
) -> None:
    try:
        with open(os.devnull, "w", encoding="utf-8") as sink:
            with redirect_stdout(sink), redirect_stderr(sink):
                payload = target(source_path)
        result_queue.put(
            {
                "status": "COMPLETE",
                "payload": payload,
                "peak_memory_bytes": _worker_peak_memory_bytes(),
            }
        )
    except MemoryError:
        result_queue.put(
            {
                "status": "RESOURCE_LIMIT",
                "payload": None,
                "peak_memory_bytes": _worker_peak_memory_bytes(),
                "failure_reason_code": "PARSER_MEMORY_LIMIT",
                "failure_exception_type": "MemoryError",
            }
        )
    except ModuleNotFoundError as error:
        result_queue.put(
            {
                "status": "UNAVAILABLE",
                "payload": None,
                "peak_memory_bytes": _worker_peak_memory_bytes(),
                "failure_reason_code": "PARSER_DEPENDENCY_UNAVAILABLE",
                "failure_exception_type": _safe_exception_type(error),
            }
        )
    except Exception as error:
        malformed = isinstance(
            error,
            (
                OSError,
                zipfile.BadZipFile,
                ElementTree.ParseError,
                _RecognizedDocumentMalformedError,
            ),
        ) or type(error).__name__ in {
            "FileDataError",
            "EmptyFileError",
            "FormatError",
            "InvalidFileException",
            "XMLSyntaxError",
        }
        result_queue.put(
            {
                "status": "MALFORMED" if malformed else "PARSER_FAILED",
                "payload": None,
                "peak_memory_bytes": _worker_peak_memory_bytes(),
                "failure_reason_code": (
                    "PARSER_INPUT_MALFORMED" if malformed else "PARSER_BACKEND_EXCEPTION"
                ),
                "failure_exception_type": _safe_exception_type(error),
            }
        )


@dataclass(slots=True)
class IsolatedParserWorker:
    """One owned subprocess invocation with an enforced deadline and RSS bound."""

    worker_target: WorkerTarget
    timeout_seconds: float = MAX_PARSER_ELAPSED_SECONDS
    memory_bytes: int = MAX_PARSER_MEMORY_BYTES

    def run(self, source_path: Path) -> _WorkerExecution:
        context = multiprocessing.get_context("spawn")
        result_queue: multiprocessing.queues.Queue[Any] = context.Queue(maxsize=1)
        process = context.Process(
            target=_worker_entry,
            args=(self.worker_target, str(source_path), result_queue),
        )
        started = monotonic()
        process.start()
        peak_memory: int | None = None
        exceeded_memory = False
        message: object | None = None
        while process.is_alive():
            if message is None:
                try:
                    message = result_queue.get_nowait()
                except queue.Empty:
                    pass
            elapsed = monotonic() - started
            if elapsed >= self.timeout_seconds:
                process.terminate()
                process.join()
                result_queue.close()
                return _WorkerExecution(
                    "TIMEOUT",
                    None,
                    int(elapsed * 1_000),
                    peak_memory,
                    "PARSER_TIMEOUT",
                )
            try:
                rss = psutil.Process(process.pid).memory_info().rss
                peak_memory = max(peak_memory or 0, rss)
                if rss > self.memory_bytes:
                    exceeded_memory = True
                    process.terminate()
                    process.join()
                    break
            except (psutil.Error, ProcessLookupError):
                pass
            process.join(min(0.05, self.timeout_seconds - elapsed))
        elapsed_ms = int((monotonic() - started) * 1_000)
        if exceeded_memory:
            result_queue.close()
            return _WorkerExecution(
                "RESOURCE_LIMIT",
                None,
                elapsed_ms,
                peak_memory,
                "PARSER_MEMORY_LIMIT",
            )
        if message is None:
            try:
                message = result_queue.get(timeout=0.2)
            except queue.Empty:
                message = None
        result_queue.close()
        if not isinstance(message, dict):
            return _WorkerExecution(
                "PARSER_FAILED", None, elapsed_ms, None, "PARSER_PROCESS_NO_RESULT"
            )
        status = message.get("status")
        worker_peak_memory = message.get("peak_memory_bytes")
        if not isinstance(status, str) or (
            worker_peak_memory is not None and not isinstance(worker_peak_memory, int)
        ):
            return _WorkerExecution(
                "PARSER_FAILED", None, elapsed_ms, None, "PARSER_PROTOCOL_INVALID"
            )
        payload = message.get("payload")
        failure_reason_code = message.get("failure_reason_code")
        failure_exception_type = message.get("failure_exception_type")
        if (failure_reason_code is not None and not isinstance(failure_reason_code, str)) or (
            failure_exception_type is not None and not isinstance(failure_exception_type, str)
        ):
            return _WorkerExecution(
                "PARSER_FAILED", None, elapsed_ms, None, "PARSER_PROTOCOL_INVALID"
            )
        if worker_peak_memory is not None:
            peak_memory = max(peak_memory or 0, worker_peak_memory)
        return _WorkerExecution(
            status,
            payload if isinstance(payload, dict) else None,
            elapsed_ms,
            peak_memory,
            failure_reason_code,
            failure_exception_type,
        )


@dataclass(slots=True)
class IsolatedPdfWorker(IsolatedParserWorker):
    """The existing PDF specialization over the shared owned-worker model."""

    worker_target: WorkerTarget = None  # type: ignore[assignment]
    timeout_seconds: float = MAX_PDF_PARSER_ELAPSED_SECONDS
    memory_bytes: int = MAX_PDF_PARSER_MEMORY_BYTES

    def __post_init__(self) -> None:
        if self.worker_target is None:
            self.worker_target = _pymupdf4llm_pdf_worker


@dataclass(slots=True)
class IsolatedXlsxWorker(IsolatedParserWorker):
    """The XLSX specialization over the same owned-worker model."""

    worker_target: WorkerTarget = None  # type: ignore[assignment]

    def __post_init__(self) -> None:
        if self.worker_target is None:
            self.worker_target = _openpyxl_xlsx_worker


@dataclass(slots=True)
class IsolatedPptxWorker(IsolatedParserWorker):
    """The PPTX specialization over the same owned-worker model."""

    worker_target: WorkerTarget = None  # type: ignore[assignment]

    def __post_init__(self) -> None:
        if self.worker_target is None:
            self.worker_target = _python_pptx_worker


@dataclass(slots=True)
class IsolatedDocxWorker(IsolatedParserWorker):
    """The restricted DOCX specialization over the shared owned-worker model."""

    worker_target: WorkerTarget = None  # type: ignore[assignment]

    def __post_init__(self) -> None:
        if self.worker_target is None:
            self.worker_target = _markitdown_docx_worker


@dataclass(slots=True)
class IsolatedDoclingWorker(IsolatedParserWorker):
    """The unified deep-parser specialization over the owned-worker model."""

    worker_target: WorkerTarget = None  # type: ignore[assignment]
    timeout_seconds: float = MAX_DEEP_PARSER_ELAPSED_SECONDS
    memory_bytes: int = MAX_DEEP_PARSER_MEMORY_BYTES

    def __post_init__(self) -> None:
        if self.worker_target is None:
            self.worker_target = docling_document_worker


@dataclass(slots=True)
class IsolatedEnrichedDoclingWorker(IsolatedDoclingWorker):
    """The explicit formula/code/picture enrichment profile."""

    def __post_init__(self) -> None:
        if self.worker_target is None:
            self.worker_target = docling_enriched_document_worker


@dataclass(slots=True)
class IsolatedMacOcrWorker(IsolatedDoclingWorker):
    """The local macOS Vision fallback for low-quality OCR results."""

    def __post_init__(self) -> None:
        if self.worker_target is None:
            self.worker_target = docling_macos_ocr_worker


def _adaptive_worker_budget(
    worker: IsolatedParserWorker,
    *,
    source_bytes: int,
    expanded_bytes: int,
) -> IsolatedParserWorker:
    """Scale elapsed work by admitted bytes while retaining a fixed RSS ceiling."""
    work_bytes = max(source_bytes, expanded_bytes)
    if work_bytes <= STREAMING_QUERY_MAP_THRESHOLD_BYTES:
        return worker
    additional_units = math.ceil(
        (work_bytes - STREAMING_QUERY_MAP_THRESHOLD_BYTES) / (64 * 1024 * 1024)
    )
    timeout = min(
        MAX_ADAPTIVE_PARSER_ELAPSED_SECONDS,
        worker.timeout_seconds + (additional_units * 15.0),
    )
    return replace(worker, timeout_seconds=timeout)


def _pdf_native_text_probe(source_path: Path) -> tuple[int, int] | None:
    """Return page and low-native-text counts without rendering page images."""
    try:
        pymupdf = import_module("pymupdf")
        document = pymupdf.open(source_path)
    except (ImportError, OSError, RuntimeError, ValueError):
        return None
    try:
        page_count = int(document.page_count)
        low_text_pages = 0
        for page_index in range(page_count):
            page = document.load_page(page_index)
            text = page.get_text("text")
            if sum(character.isalnum() for character in text) < MIN_PDF_NATIVE_TEXT_CHARACTERS:
                low_text_pages += 1
        return page_count, low_text_pages
    finally:
        document.close()


def _pdf_page_count_for_query_routing(source_path: Path) -> int | None:
    """Read the PDF page tree and native-text coverage for query routing."""
    probe = _pdf_native_text_probe(source_path)
    return None if probe is None else probe[0]


def _streaming_query_excerpt(text: str, query: str) -> str | None:
    match = match_document_text(text, query)
    if match is None:
        return None
    radius = MAX_STREAMING_QUERY_EXCERPT_CHARS // 2
    start = max(0, match.source_start - radius)
    end = min(len(text), max(match.source_end, match.source_start) + radius)
    return text[start:end][:MAX_STREAMING_QUERY_EXCERPT_CHARS]


def _pdf_ocr_render_scale(page: object) -> float:
    """Bound one source-relative page render without raising a global size ceiling."""
    rect = getattr(page, "rect", None)
    width = float(getattr(rect, "width", 0.0))
    height = float(getattr(rect, "height", 0.0))
    if width <= 0 or height <= 0:
        raise RuntimeError("PDF page has invalid render dimensions")
    native_pixels = width * height
    if native_pixels <= MAX_PDF_PAGE_OCR_PIXELS:
        return PDF_PAGE_OCR_RENDER_SCALE
    return max(0.1, math.sqrt(MAX_PDF_PAGE_OCR_PIXELS / native_pixels))


def _rapidocr_page_lines(
    page: Any,
    engine: Callable[[bytes], object],
    *,
    page_number: int,
) -> tuple[list[dict[str, object]], float]:
    """OCR one ephemeral raster and retain only bounded text/region facts."""
    pymupdf = import_module("pymupdf")
    scale = _pdf_ocr_render_scale(page)
    pixmap = page.get_pixmap(
        matrix=pymupdf.Matrix(scale, scale),
        alpha=False,
        colorspace=pymupdf.csRGB,
    )
    width = int(pixmap.width)
    height = int(pixmap.height)
    if width <= 0 or height <= 0 or width * height > MAX_PDF_PAGE_OCR_PIXELS:
        raise MemoryError("PDF OCR render exceeds its pixel bound")
    image_bytes = pixmap.tobytes("png")
    del pixmap
    output = engine(image_bytes)
    del image_bytes
    texts = getattr(output, "txts", None)
    scores = getattr(output, "scores", None)
    boxes = getattr(output, "boxes", None)
    if texts is None:
        del output
        gc.collect()
        return [], scale
    if not isinstance(texts, (list, tuple)):
        raise RuntimeError("RapidOCR returned invalid text output")
    lines: list[dict[str, object]] = []
    for line_index, text in enumerate(texts, start=1):
        if not isinstance(text, str) or not text.strip():
            continue
        extension: dict[str, object] = {
            "text_source": "LOCAL_OCR",
            "ocr_engine": "RapidOCR",
            "ocr_render_scale": round(scale, 6),
        }
        if isinstance(scores, (list, tuple)) and line_index <= len(scores):
            score = scores[line_index - 1]
            if isinstance(score, (int, float)) and not isinstance(score, bool):
                extension["ocr_confidence"] = round(float(score), 6)
        if boxes is not None:
            try:
                raw_box = boxes[line_index - 1]
                points = raw_box.tolist() if hasattr(raw_box, "tolist") else raw_box
                coordinates = [
                    (float(point[0]), float(point[1]))
                    for point in points
                    if isinstance(point, (list, tuple)) and len(point) >= 2
                ]
                if coordinates:
                    left = min(point[0] for point in coordinates)
                    top = min(point[1] for point in coordinates)
                    right = max(point[0] for point in coordinates)
                    bottom = max(point[1] for point in coordinates)
                    if 0 <= left < right <= width and 0 <= top < bottom <= height:
                        normalized_region = [
                            round(left / width, 6),
                            round(top / height, 6),
                            round(right / width, 6),
                            round(bottom / height, 6),
                        ]
                        extension["normalized_region"] = normalized_region
                        page_rect = page.rect
                        page_width = float(page_rect.width)
                        page_height = float(page_rect.height)
                        extension["visual_region"] = {
                            "page": page_number,
                            "bbox": [
                                round(normalized_region[0] * page_width, 6),
                                round(normalized_region[1] * page_height, 6),
                                round(normalized_region[2] * page_width, 6),
                                round(normalized_region[3] * page_height, 6),
                            ],
                            "page_size": [page_width, page_height],
                            "coordinate_space": "PAGE_POINTS_TOP_LEFT",
                        }
            except (IndexError, TypeError, ValueError):
                pass
        lines.append(
            {
                "text": text.strip(),
                "line": line_index,
                "extension": extension,
            }
        )
    del output
    gc.collect()
    return lines, scale


def _rapidocr_engine() -> Callable[[bytes], object]:
    module = import_module("rapidocr")
    engine_type = getattr(module, "RapidOCR", None)
    if not callable(engine_type):
        raise ModuleNotFoundError("RapidOCR runtime is unavailable")
    return cast(Callable[[bytes], object], engine_type())


def _pdf_page_text_projection(
    source_path: str,
    *,
    query: str | None,
) -> tuple[list[dict[str, object]], int, int, int, list[str]]:
    """Project native text and page-local OCR without retaining source rasters."""
    pymupdf = import_module("pymupdf")
    document = pymupdf.open(source_path)
    items: list[dict[str, object]] = []
    ocr_page_count = 0
    native_text_page_count = 0
    text_characters = 0
    engine: Callable[[bytes], object] | None = None
    ocr_render_scales: set[float] = set()
    try:
        for page_index in range(document.page_count):
            page = document.load_page(page_index)
            text = page.get_text("text")
            is_low_text = (
                sum(character.isalnum() for character in text)
                < MIN_PDF_NATIVE_TEXT_CHARACTERS
            )
            ocr_lines: list[dict[str, object]] = []
            if is_low_text:
                if engine is None:
                    engine = _rapidocr_engine()
                ocr_lines, scale = _rapidocr_page_lines(
                    page,
                    engine,
                    page_number=page_index + 1,
                )
                ocr_render_scales.add(round(scale, 6))
                text = "\n".join(str(line["text"]) for line in ocr_lines)
                ocr_page_count += 1
            else:
                native_text_page_count += 1
            text_characters += len(text)
            if query is not None:
                excerpt = _streaming_query_excerpt(text, query)
                if excerpt is None:
                    continue
                if len(items) >= MAX_STREAMING_QUERY_MATCH_ITEMS:
                    raise MemoryError("streaming PDF query map exceeds its result bound")
                items.append(
                    {
                        "kind": "pdf_ocr_page_block" if is_low_text else "pdf_page_block",
                        "role": "SECTION",
                        "text_or_value": excerpt,
                        "parent": None,
                        "location": {"page": page_index + 1, "block": 1},
                        "extension": {
                            "page": page_index + 1,
                            "region_kind": "query_excerpt",
                            "text_source": "LOCAL_OCR" if is_low_text else "EMBEDDED_TEXT",
                            **(
                                {
                                    "ocr_engine": "RapidOCR",
                                    "ocr_line_count": len(ocr_lines),
                                }
                                if is_low_text
                                else {}
                            ),
                        },
                    }
                )
                continue
            if is_low_text:
                for line in ocr_lines:
                    if len(items) >= MAX_PARSED_ITEMS_OR_BLOCKS:
                        raise MemoryError("PDF page OCR exceeds its item bound")
                    items.append(
                        {
                            "node_id": (
                                f"pdf:page:{page_index + 1}:ocr-line:{line['line']}"
                            ),
                            "kind": "pdf_ocr_text_line",
                            "role": "PARAGRAPH",
                            "text_or_value": line["text"],
                            "parent": None,
                            "location": {
                                "page": page_index + 1,
                                "line": line["line"],
                            },
                            "extension": line["extension"],
                        }
                    )
            elif text.strip():
                items.append(
                    {
                        "kind": "pdf_page_block",
                        "role": "SECTION",
                        "text_or_value": text,
                        "parent": None,
                        "location": {"page": page_index + 1, "block": 1},
                        "extension": {
                            "page": page_index + 1,
                            "region_kind": "page",
                            "text_source": "EMBEDDED_TEXT",
                        },
                    }
                )
        page_count = document.page_count
    finally:
        document.close()
    warnings = [
        f"PDF_PAGE_TEXT_PAGES:{page_count}",
        f"PDF_PAGE_TEXT_CHARACTERS:{text_characters}",
        f"PDF_PAGE_EMBEDDED_TEXT_PAGES:{native_text_page_count}",
        f"PDF_PAGE_OCR_PAGES:{ocr_page_count}",
    ]
    if ocr_page_count:
        render_scale_warning = (
            f"OCR_RENDER_SCALE:{next(iter(ocr_render_scales))}"
            if len(ocr_render_scales) == 1
            else "OCR_RENDER_SCALES:"
            + ",".join(str(value) for value in sorted(ocr_render_scales))
        )
        warnings.extend(
            (
                "OCR_ENGINE:RAPIDOCR_ONNXRUNTIME",
                render_scale_warning,
                "OCR_TEXT_AUTHORITY:MODEL_DERIVED",
            )
        )
    return items, page_count, native_text_page_count, ocr_page_count, warnings


def _streaming_pdf_query_map(source_path: str, query: str) -> dict[str, Any]:
    """Scan PDF pages independently with bounded page-local OCR where needed."""
    items, page_count, _native_pages, ocr_page_count, page_warnings = (
        _pdf_page_text_projection(source_path, query=query)
    )
    native_items, native_warnings = project_pdf_native(source_path)
    native_match_count = 0
    for native_item in native_items:
        excerpt = _streaming_query_excerpt(str(native_item.get("text_or_value") or ""), query)
        if excerpt is None:
            continue
        if len(items) >= MAX_STREAMING_QUERY_MATCH_ITEMS:
            raise MemoryError("streaming PDF query map exceeds its result bound")
        items.append({**native_item, "text_or_value": excerpt})
        native_match_count += 1
    return {
        "backend_name": "STEWARDStreamingMap",
        "backend_version": version("local-system-steward"),
        "warnings": [
            f"STREAMING_QUERY_MAP_PAGES:{page_count}",
            f"STREAMING_QUERY_MAP_OCR_PAGES:{ocr_page_count}",
            f"STREAMING_QUERY_MAP_NATIVE_MATCHES:{native_match_count}",
            *page_warnings,
            *native_warnings,
        ],
        "items": items,
        **(
            {
                "resource_extension": {
                    "ocr_backend": "RapidOCR",
                    "ocr_version": version("rapidocr"),
                    "ocr_page_count": ocr_page_count,
                    "ocr_text_authority": "MODEL_DERIVED",
                }
            }
            if ocr_page_count
            else {}
        ),
    }


def _rapidocr_pdf_worker(source_path: str) -> dict[str, Any]:
    """Read scan-heavy PDFs with one bounded local OCR engine and page disposal."""
    items, _pages, _native_pages, _ocr_pages, warnings = _pdf_page_text_projection(
        source_path,
        query=None,
    )
    native_items, native_warnings = project_pdf_native(
        source_path,
        include_page_auxiliary=False,
    )
    return {
        "backend_name": "STEWARDPageOCR",
        "backend_version": version("rapidocr"),
        "warnings": [*warnings, *native_warnings],
        "items": [*items, *native_items],
        "resource_extension": {
            "ocr_backend": "RapidOCR",
            "ocr_version": version("rapidocr"),
            "ocr_page_count": _ocr_pages,
            "ocr_text_authority": "MODEL_DERIVED",
        },
    }


def _streaming_xlsx_shared_string_cells(
    archive: zipfile.ZipFile,
    member: str,
    *,
    sheet_name: str,
    sheet_index: int,
    shared_matches: dict[int, str],
    items: list[dict[str, object]],
) -> tuple[int, bool]:
    """Resolve matched shared-string indexes without per-cell Python callbacks.

    Large worksheets may contain millions of cells.  Their shared-string table
    already supplies the normalized query decision, so this path streams the
    compressed worksheet once and asks the C regex engine only for the bounded
    set of matching indexes.  Individual matching cell fragments are then
    validated before publication.
    """

    if not shared_matches or len(shared_matches) > 128:
        return 0, False
    alternatives = b"|".join(
        re.escape(str(index).encode("ascii")) for index in sorted(shared_matches)
    )
    value_pattern = re.compile(rb"<v>\s*(" + alternatives + rb")\s*</v>")
    coordinate_pattern = re.compile(rb'\br="([^"]{1,64})"')
    buffer = b""
    candidate_cells = 0
    with archive.open(member) as source:
        while chunk := source.read(DOCUMENT_INGRESS_CHUNK_BYTES):
            buffer += chunk
            complete_end = buffer.rfind(b"</c>")
            if complete_end < 0:
                if len(buffer) > MAX_CONTROL_XML_BYTES:
                    raise MemoryError("streaming XLSX cell fragment exceeds its bound")
                continue
            complete_end += len(b"</c>")
            complete = buffer[:complete_end]
            buffer = buffer[complete_end:]
            for match in value_pattern.finditer(complete):
                cell_start = complete.rfind(b"<c", 0, match.start())
                cell_end = complete.find(b"</c>", match.end())
                if cell_start < 0 or cell_end < 0:
                    continue
                start_tag_end = complete.find(b">", cell_start, match.start())
                if start_tag_end < 0:
                    continue
                start_tag = complete[cell_start : start_tag_end + 1]
                if not re.search(rb'\bt="s"', start_tag):
                    continue
                coordinate_match = coordinate_pattern.search(start_tag)
                if coordinate_match is None:
                    continue
                candidate_cells += 1
                index = int(match.group(1))
                excerpt = shared_matches.get(index)
                if excerpt is None:
                    continue
                items.append(
                    {
                        "kind": "xlsx_cell",
                        "role": "TABLE_CELL",
                        "text_or_value": excerpt,
                        "parent": f"sheet:{sheet_index}",
                        "location": {
                            "sheet": sheet_name,
                            "sheet_index": sheet_index,
                            "cell": coordinate_match.group(1).decode("utf-8", "strict"),
                        },
                        "extension": {
                            "streaming_query_map": True,
                            "shared_string_index": index,
                        },
                    }
                )
                if len(items) >= MAX_STREAMING_QUERY_MATCH_ITEMS:
                    return candidate_cells, True
    return candidate_cells, False


def _streaming_xlsx_query_map(source_path: str, query: str) -> dict[str, Any]:
    """Scan native worksheet XML and retain only matching cells.

    This avoids openpyxl's per-cell object construction for very large sparse
    workbooks while preserving sheet and A1-coordinate provenance.
    """
    items: list[dict[str, object]] = []
    shared_matches: dict[int, str] = {}
    scanned_shared_strings = 0
    scanned_cells = 0
    early_stop = False
    indexed_shared_string_map = False
    component_warnings: list[str] = []
    with zipfile.ZipFile(source_path) as archive:
        names = set(archive.namelist())
        if "xl/sharedStrings.xml" in names:
            in_shared_item = False
            in_shared_text = False
            shared_fragments: list[str] = []

            def shared_start(name: str, _attrs: dict[str, str]) -> None:
                nonlocal in_shared_item, in_shared_text, shared_fragments
                local_name = _xml_local_name(name)
                if local_name == "si":
                    in_shared_item = True
                    shared_fragments = []
                elif in_shared_item and local_name == "t":
                    in_shared_text = True

            def shared_data(value: str) -> None:
                if in_shared_text:
                    shared_fragments.append(value)

            def shared_end(name: str) -> None:
                nonlocal in_shared_item, in_shared_text, scanned_shared_strings
                local_name = _xml_local_name(name)
                if local_name == "t":
                    in_shared_text = False
                elif local_name == "si":
                    excerpt = _streaming_query_excerpt("".join(shared_fragments), query)
                    if excerpt is not None:
                        shared_matches[scanned_shared_strings] = excerpt
                    scanned_shared_strings += 1
                    in_shared_item = False

            shared_parser = expat.ParserCreate(namespace_separator="}")
            shared_parser.StartElementHandler = shared_start
            shared_parser.CharacterDataHandler = shared_data
            shared_parser.EndElementHandler = shared_end
            with archive.open("xl/sharedStrings.xml") as source:
                while chunk := source.read(DOCUMENT_INGRESS_CHUNK_BYTES):
                    shared_parser.Parse(chunk, False)
                shared_parser.Parse(b"", True)

        relationship_targets: dict[str, str] = {}
        with archive.open("xl/_rels/workbook.xml.rels") as source:
            for _event, element in ElementTree.iterparse(source, events=("end",)):
                if _xml_local_name(element.tag) != "Relationship":
                    continue
                identifier = element.attrib.get("Id")
                target = element.attrib.get("Target")
                if identifier and target and element.attrib.get("TargetMode") != "External":
                    normalized = target.lstrip("/")
                    if not normalized.startswith("xl/"):
                        normalized = f"xl/{normalized}"
                    relationship_targets[identifier] = PurePosixPath(normalized).as_posix()
                element.clear()

        sheet_records: list[tuple[str, str]] = []
        with archive.open("xl/workbook.xml") as source:
            for _event, element in ElementTree.iterparse(source, events=("end",)):
                if _xml_local_name(element.tag) != "sheet":
                    continue
                name = element.attrib.get("name")
                relationship_id = next(
                    (
                        value
                        for key, value in element.attrib.items()
                        if _xml_local_name(key) == "id"
                    ),
                    None,
                )
                target = relationship_targets.get(relationship_id or "")
                if name and target and target in names:
                    sheet_records.append((name, target))
                element.clear()

        for sheet_index, (sheet_name, target) in enumerate(sheet_records, start=1):
            if 0 < len(shared_matches) <= 128:
                indexed_shared_string_map = True
                candidate_cells, early_stop = _streaming_xlsx_shared_string_cells(
                    archive,
                    target,
                    sheet_name=sheet_name,
                    sheet_index=sheet_index,
                    shared_matches=shared_matches,
                    items=items,
                )
                scanned_cells += candidate_cells
            else:
                current_cell: dict[str, object] | None = None
                active_text: str | None = None

                def sheet_start(name: str, attrs: dict[str, str]) -> None:
                    nonlocal current_cell, active_text
                    local_name = _xml_local_name(name)
                    if local_name == "c":
                        current_cell = {
                            "type": attrs.get("t"),
                            "coordinate": attrs.get("r"),
                            "formula": [],
                            "value": [],
                            "inline": [],
                        }
                    elif current_cell is not None and local_name in {"f", "v", "t"}:
                        active_text = local_name

                def sheet_data(value: str) -> None:
                    if current_cell is not None and active_text is not None:
                        key = (
                            "inline"
                            if active_text == "t"
                            else ("formula" if active_text == "f" else "value")
                        )
                        fragments = current_cell[key]
                        assert isinstance(fragments, list)
                        fragments.append(value)

                def sheet_end(name: str) -> None:
                    nonlocal current_cell, active_text, scanned_cells, early_stop
                    local_name = _xml_local_name(name)
                    if local_name in {"f", "v", "t"}:
                        active_text = None
                        return
                    if local_name != "c" or current_cell is None:
                        return
                    scanned_cells += 1
                    cell_type = current_cell["type"]
                    coordinate = current_cell["coordinate"] or f"cell-{scanned_cells}"
                    formula_text = "".join(current_cell["formula"])  # type: ignore[arg-type]
                    value_text = "".join(current_cell["value"])  # type: ignore[arg-type]
                    inline_text = "".join(current_cell["inline"])  # type: ignore[arg-type]
                    formula = f"={formula_text}" if formula_text else None
                    excerpt: str | None = None
                    if cell_type == "s":
                        try:
                            excerpt = shared_matches.get(int(value_text or "-1"))
                        except ValueError:
                            excerpt = None
                    elif cell_type == "inlineStr":
                        excerpt = _streaming_query_excerpt(inline_text, query)
                    elif formula is not None:
                        excerpt = _streaming_query_excerpt(formula, query)
                    elif cell_type == "str":
                        excerpt = _streaming_query_excerpt(value_text, query)
                    elif value_text:
                        excerpt = _streaming_query_excerpt(value_text, query)
                    if excerpt is not None and len(items) < MAX_STREAMING_QUERY_MATCH_ITEMS:
                        items.append(
                            {
                                "kind": "xlsx_cell",
                                "role": "FORMULA" if formula is not None else "TABLE_CELL",
                                "text_or_value": excerpt,
                                "parent": f"sheet:{sheet_index}",
                                "location": {
                                    "sheet": sheet_name,
                                    "sheet_index": sheet_index,
                                    "cell": str(coordinate),
                                },
                                "extension": {
                                    **({"formula": formula} if formula is not None else {}),
                                    "streaming_query_map": True,
                                },
                            }
                        )
                    if len(items) >= MAX_STREAMING_QUERY_MATCH_ITEMS:
                        early_stop = True
                    current_cell = None

                sheet_parser = expat.ParserCreate(namespace_separator="}")
                sheet_parser.StartElementHandler = sheet_start
                sheet_parser.CharacterDataHandler = sheet_data
                sheet_parser.EndElementHandler = sheet_end
                with archive.open(target) as source:
                    while chunk := source.read(DOCUMENT_INGRESS_CHUNK_BYTES):
                        sheet_parser.Parse(chunk, False)
                        if early_stop:
                            break
                    if not early_stop:
                        sheet_parser.Parse(b"", True)
            if early_stop:
                break
            relationship_name = (
                f"{PurePosixPath(target).parent.as_posix()}/_rels/{PurePosixPath(target).name}.rels"
            )
            relationships = _internal_relationship_targets(archive, relationship_name, target)
            for relationship_kind, related_target in relationships:
                if related_target not in names:
                    continue
                if relationship_kind == "comments":
                    comments_root = _streaming_optional_xml_root(
                        archive,
                        related_target,
                        component=f"comments:sheet:{sheet_index}",
                        warnings=component_warnings,
                    )
                    if comments_root is None:
                        continue
                    authors = [
                        _xml_element_text(author)
                        for author in comments_root.iter()
                        if _xml_local_name(author.tag) == "author"
                    ]
                    for comment in comments_root.iter():
                        if _xml_local_name(comment.tag) != "comment":
                            continue
                        reference = comment.attrib.get("ref", "comment")
                        author: str | None = None
                        try:
                            author = authors[int(comment.attrib.get("authorId", "-1"))]
                        except (IndexError, ValueError):
                            author = None
                        excerpt = _streaming_query_excerpt(_xml_element_text(comment), query)
                        if excerpt is None:
                            continue
                        if len(items) >= MAX_STREAMING_QUERY_MATCH_ITEMS:
                            early_stop = True
                            break
                        items.append(
                            {
                                "kind": "xlsx_comment",
                                "role": "NOTE",
                                "text_or_value": excerpt,
                                "parent": f"sheet:{sheet_index}",
                                "location": {
                                    "sheet": sheet_name,
                                    "sheet_index": sheet_index,
                                    "cell": reference,
                                    "comment": 1,
                                },
                                "extension": {
                                    "streaming_query_map": True,
                                    **({"author": author[:256]} if author else {}),
                                },
                            }
                        )
                elif relationship_kind == "drawing":
                    drawing_relationship_name = (
                        f"{PurePosixPath(related_target).parent.as_posix()}/_rels/"
                        f"{PurePosixPath(related_target).name}.rels"
                    )
                    drawing_relationships = _internal_relationship_targets(
                        archive, drawing_relationship_name, related_target
                    )
                    chart_index = 0
                    for drawing_kind, chart_target in drawing_relationships:
                        if drawing_kind != "chart" or chart_target not in names:
                            continue
                        chart_index += 1
                        chart_root = _streaming_optional_xml_root(
                            archive,
                            chart_target,
                            component=f"chart:sheet:{sheet_index}:{chart_index}",
                            warnings=component_warnings,
                        )
                        if chart_root is None:
                            continue
                        _append_streaming_match(
                            items,
                            text=_xml_element_text(chart_root),
                            query=query,
                            kind="xlsx_chart",
                            role="FIGURE",
                            parent=f"sheet:{sheet_index}",
                            location={
                                "sheet": sheet_name,
                                "sheet_index": sheet_index,
                                "chart": chart_index,
                            },
                        )
                if early_stop:
                    break
            if early_stop:
                break
    return {
        "backend_name": "STEWARDStreamingMap",
        "backend_version": version("local-system-steward"),
        "warnings": [
            f"STREAMING_QUERY_MAP_SHEETS:{len(sheet_records)}",
            f"STREAMING_QUERY_MAP_SHARED_STRINGS:{scanned_shared_strings}",
            f"STREAMING_QUERY_MAP_CELLS_SCANNED:{scanned_cells}",
            *(["STREAMING_QUERY_MAP_INDEXED_SHARED_STRINGS"] if indexed_shared_string_map else []),
            *component_warnings,
            *(
                [f"STREAMING_QUERY_MAP_EARLY_STOP_LIMIT:{MAX_STREAMING_QUERY_MATCH_ITEMS}"]
                if early_stop
                else []
            ),
        ],
        "items": items,
    }


def _xml_local_name(tag: object) -> str:
    return tag.rsplit("}", 1)[-1] if isinstance(tag, str) else ""


def _xml_element_text(element: ElementTree.Element) -> str:
    return "".join(
        value for node in element.iter() for value in (node.text,) if isinstance(value, str)
    ).strip()


def _internal_relationship_targets(
    archive: zipfile.ZipFile,
    relationships_member: str,
    owner_member: str,
) -> list[tuple[str, str]]:
    """Resolve internal OPC relationships relative to their owning part."""

    if relationships_member not in archive.namelist():
        return []
    targets: list[tuple[str, str]] = []
    owner_parent = PurePosixPath(owner_member).parent.as_posix()
    with archive.open(relationships_member) as source:
        for _event, element in ElementTree.iterparse(source, events=("end",)):
            if _xml_local_name(element.tag) != "Relationship":
                continue
            if element.attrib.get("TargetMode") == "External":
                element.clear()
                continue
            relationship_type = element.attrib.get("Type", "")
            target = element.attrib.get("Target")
            if target:
                if target.startswith("/"):
                    normalized = target.lstrip("/")
                else:
                    normalized = posixpath.normpath(PurePosixPath(owner_parent, target).as_posix())
                targets.append((relationship_type.rsplit("/", 1)[-1], normalized))
            element.clear()
    return targets


def _streaming_optional_xml_root(
    archive: zipfile.ZipFile,
    member: str,
    *,
    component: str,
    warnings: list[str],
) -> ElementTree.Element | None:
    """Read one admitted optional XML part while preserving other valid parts."""

    try:
        info = archive.getinfo(member)
        if info.file_size > MAX_OPTIONAL_XML_BYTES:
            warnings.append(f"STREAMING_COMPONENT_RESOURCE_LIMIT:{component}")
            return None
        with archive.open(info) as source:
            payload = source.read(MAX_OPTIONAL_XML_BYTES + 1)
        if len(payload) > MAX_OPTIONAL_XML_BYTES:
            warnings.append(f"STREAMING_COMPONENT_RESOURCE_LIMIT:{component}")
            return None
        return ElementTree.fromstring(payload)
    except (ElementTree.ParseError, KeyError, OSError, RuntimeError):
        warnings.append(f"STREAMING_COMPONENT_MALFORMED:{component}")
        return None


def _append_streaming_match(
    items: list[dict[str, object]],
    *,
    text: str,
    query: str,
    kind: str,
    role: str,
    parent: str | None,
    location: dict[str, int | str],
) -> None:
    excerpt = _streaming_query_excerpt(text, query)
    if excerpt is None:
        return
    if len(items) >= MAX_STREAMING_QUERY_MATCH_ITEMS:
        raise MemoryError("streaming package query map exceeds its result bound")
    items.append(
        {
            "kind": kind,
            "role": role,
            "text_or_value": excerpt,
            "parent": parent,
            "location": location,
            "extension": {"streaming_query_map": True},
        }
    )


def _streaming_docx_query_map(
    archive: zipfile.ZipFile, query: str
) -> tuple[list[dict[str, object]], int, list[str]]:
    items: list[dict[str, object]] = []
    warnings: list[str] = []
    block = 0
    with archive.open("word/document.xml") as source:
        for _event, element in ElementTree.iterparse(source, events=("end",)):
            if _xml_local_name(element.tag) != "p":
                continue
            block += 1
            _append_streaming_match(
                items,
                text=_xml_element_text(element),
                query=query,
                kind="docx_paragraph",
                role="PARAGRAPH",
                parent="document:current",
                location={"block": block},
            )
            element.clear()
    for kind, member, element_name, location_name in (
        ("docx_comment", "word/comments.xml", "comment", "comment"),
        ("docx_footnote", "word/footnotes.xml", "footnote", "footnote"),
        ("docx_endnote", "word/endnotes.xml", "endnote", "endnote"),
    ):
        if member not in archive.namelist():
            continue
        native_index = 0
        root = _streaming_optional_xml_root(
            archive, member, component=location_name, warnings=warnings
        )
        if root is None:
            continue
        for element in root.iter():
            if _xml_local_name(element.tag) != element_name:
                continue
            identifier = next(
                (value for key, value in element.attrib.items() if _xml_local_name(key) == "id"),
                None,
            )
            if identifier is not None and identifier.startswith("-"):
                continue
            native_index += 1
            _append_streaming_match(
                items,
                text=_xml_element_text(element),
                query=query,
                kind=kind,
                role="NOTE",
                parent="document:current",
                location={location_name: identifier or native_index},
            )
    return items, block, warnings


def _streaming_pptx_query_map(
    archive: zipfile.ZipFile, query: str
) -> tuple[list[dict[str, object]], int, list[str]]:
    items: list[dict[str, object]] = []
    warnings: list[str] = []
    archive_names = set(archive.namelist())
    slide_names = sorted(
        (name for name in archive_names if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)),
        key=lambda name: int(Path(name).stem.removeprefix("slide")),
    )
    for slide_index, name in enumerate(slide_names, start=1):
        paragraph = 0
        with archive.open(name) as source:
            for _event, element in ElementTree.iterparse(source, events=("end",)):
                if _xml_local_name(element.tag) != "p":
                    continue
                paragraph += 1
                _append_streaming_match(
                    items,
                    text=_xml_element_text(element),
                    query=query,
                    kind="pptx_text",
                    role="PARAGRAPH",
                    parent=f"slide:{slide_index}",
                    location={"slide": slide_index, "paragraph": paragraph},
                )
                element.clear()
        relationship_name = f"ppt/slides/_rels/{Path(name).name}.rels"
        relationships = _internal_relationship_targets(archive, relationship_name, name)
        notes_names = [target for kind, target in relationships if kind == "notesSlide"]
        for notes_name in notes_names:
            if notes_name not in archive_names:
                continue
            notes_root = _streaming_optional_xml_root(
                archive,
                notes_name,
                component=f"notes:slide:{slide_index}",
                warnings=warnings,
            )
            if notes_root is None:
                continue
            notes_paragraph = 0
            for element in notes_root.iter():
                if _xml_local_name(element.tag) != "p":
                    continue
                notes_paragraph += 1
                _append_streaming_match(
                    items,
                    text=_xml_element_text(element),
                    query=query,
                    kind="pptx_speaker_notes",
                    role="NOTE",
                    parent=f"slide:{slide_index}",
                    location={
                        "slide": slide_index,
                        "notes": 1,
                        "paragraph": notes_paragraph,
                    },
                )
        chart_names = [target for kind, target in relationships if kind == "chart"]
        for chart_index, chart_name in enumerate(chart_names, start=1):
            if chart_name not in archive_names:
                continue
            chart_root = _streaming_optional_xml_root(
                archive,
                chart_name,
                component=f"chart:slide:{slide_index}:{chart_index}",
                warnings=warnings,
            )
            if chart_root is None:
                continue
            _append_streaming_match(
                items,
                text=_xml_element_text(chart_root),
                query=query,
                kind="pptx_chart",
                role="FIGURE",
                parent=f"slide:{slide_index}",
                location={"slide": slide_index, "chart": chart_index},
            )
    return items, len(slide_names), warnings


class _TolerantEpubHtmlMapParser(HTMLParser):
    """Bounded HTML fallback for EPUB content that is common HTML but not XML."""

    _ACCEPTED = frozenset({"p", "h1", "h2", "h3", "h4", "h5", "h6", "li", "td", "th"})

    def __init__(
        self,
        *,
        query: str | None,
        section_index: int,
        section_title: str,
        budget: _EpubNativeBudget | None = None,
    ) -> None:
        super().__init__(convert_charrefs=True)
        self.query = query
        self.section_index = section_index
        self.section_title = section_title
        self.budget = budget
        self.items: list[dict[str, object]] = []
        self._frames: list[tuple[str, list[str]]] = []
        self._ordinal = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        del attrs
        normalized = tag.casefold()
        if normalized in self._ACCEPTED:
            self._frames.append((normalized, []))

    def handle_startendtag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        self.handle_starttag(tag, attrs)
        self.handle_endtag(tag)

    def handle_data(self, data: str) -> None:
        for _tag, fragments in self._frames:
            fragments.append(data)

    def handle_endtag(self, tag: str) -> None:
        normalized = tag.casefold()
        matching = next(
            (
                index
                for index in range(len(self._frames) - 1, -1, -1)
                if self._frames[index][0] == normalized
            ),
            None,
        )
        if matching is None:
            return
        frame_tag, fragments = self._frames[matching]
        del self._frames[matching:]
        self._emit(frame_tag, "".join(fragments))

    def finish(self) -> None:
        while self._frames:
            tag, fragments = self._frames.pop(0)
            self._emit(tag, "".join(fragments))

    def _emit(self, tag: str, text: str) -> None:
        self._ordinal += 1
        _append_epub_projection_item(
            self.items,
            text=text,
            query=self.query,
            kind="epub_section_text",
            role="HEADING" if tag.startswith("h") else "PARAGRAPH",
            parent=f"section:{self.section_index}",
            location={
                "section": self.section_index,
                "section_title": self.section_title,
                "ordinal": self._ordinal,
            },
            budget=self.budget,
        )


def _tolerant_epub_html_member(
    archive: zipfile.ZipFile,
    name: str,
    query: str | None,
    section_index: int,
    budget: _EpubNativeBudget | None = None,
) -> list[dict[str, object]]:
    parser = _TolerantEpubHtmlMapParser(
        query=query,
        section_index=section_index,
        section_title=Path(name).name,
        budget=budget,
    )
    decoder = import_module("codecs").getincrementaldecoder("utf-8")("replace")
    with archive.open(name) as source:
        while chunk := source.read(DOCUMENT_INGRESS_CHUNK_BYTES):
            parser.feed(decoder.decode(chunk))
        parser.feed(decoder.decode(b"", final=True))
    parser.close()
    parser.finish()
    return parser.items


@dataclass(slots=True)
class _EpubNativeBudget:
    structure_only: bool
    remaining_text_bytes: int = MAX_EPUB_NATIVE_READ_TEXT_BYTES
    accepted_items: int = 0
    omitted_items: int = 0


def _append_epub_projection_item(
    items: list[dict[str, object]],
    *,
    text: str,
    query: str | None,
    kind: str,
    role: str,
    parent: str | None,
    location: dict[str, int | str],
    budget: _EpubNativeBudget | None = None,
    node_id: str | None = None,
) -> None:
    normalized = " ".join(text.split())
    if not normalized:
        return
    projected = _streaming_query_excerpt(normalized, query) if query is not None else normalized
    if projected is None:
        return
    if budget is not None:
        if budget.structure_only and role not in {"HEADING", "SECTION"}:
            return
        projected_bytes = len(projected.encode("utf-8"))
        if (
            budget.accepted_items >= MAX_EPUB_NATIVE_ITEMS
            or projected_bytes > budget.remaining_text_bytes
        ):
            budget.omitted_items += 1
            return
        budget.remaining_text_bytes -= projected_bytes
        budget.accepted_items += 1
    limit = MAX_STREAMING_QUERY_MATCH_ITEMS if query is not None else MAX_PARSED_ITEMS_OR_BLOCKS
    if len(items) >= limit:
        raise MemoryError("EPUB native projection exceeds its item bound")
    item: dict[str, object] = {
        "kind": kind,
        "role": role,
        "text_or_value": projected,
        "parent": parent,
        "location": location,
        "extension": (
            {"streaming_query_map": True} if query is not None else {"native_epub_projection": True}
        ),
    }
    if node_id is not None:
        item["node_id"] = node_id
    items.append(item)


def _epub_reading_order(
    archive: zipfile.ZipFile,
) -> tuple[list[str], list[str]]:
    """Resolve manifest/spine order, falling back to deterministic HTML member order."""
    warnings: list[str] = []
    names = set(archive.namelist())
    html_names = sorted(
        name for name in names if Path(name).suffix.casefold() in {".xhtml", ".html", ".htm"}
    )
    try:
        container = ElementTree.fromstring(_bounded_archive_read(archive, "META-INF/container.xml"))
        package_path = next(
            path
            for element in container.iter()
            if element.tag.rsplit("}", 1)[-1] == "rootfile"
            and isinstance((path := element.get("full-path")), str)
            and path in names
        )
        package = ElementTree.fromstring(_bounded_archive_read(archive, package_path))
        package_dir = posixpath.dirname(package_path)
        manifest: dict[str, str] = {}
        for element in package.iter():
            if element.tag.rsplit("}", 1)[-1] != "item":
                continue
            identifier = element.get("id")
            href = element.get("href")
            media_type = element.get("media-type", "")
            if (
                not identifier
                or not href
                or media_type
                not in {
                    "application/xhtml+xml",
                    "text/html",
                }
            ):
                continue
            member = posixpath.normpath(posixpath.join(package_dir, unquote(href.split("#", 1)[0])))
            if member in names:
                manifest[identifier] = member
        ordered = [
            manifest[identifier]
            for element in package.iter()
            if element.tag.rsplit("}", 1)[-1] == "itemref"
            and isinstance((identifier := element.get("idref")), str)
            and identifier in manifest
        ]
        ordered = list(dict.fromkeys((*ordered, *html_names)))
        return ordered, warnings
    except (KeyError, StopIteration, ElementTree.ParseError, OSError):
        warnings.append("EPUB_PACKAGE_ORDER_FALLBACK")
        return html_names, warnings


def _epub_projection(
    archive: zipfile.ZipFile,
    query: str | None,
    *,
    native_view: str = "READ",
) -> tuple[list[dict[str, object]], int, list[str]]:
    items: list[dict[str, object]] = []
    content_names, warnings = _epub_reading_order(archive)
    budget = _EpubNativeBudget(structure_only=native_view == "STRUCTURE") if query is None else None
    if query is None:
        items.append(
            {
                "kind": "epub_document",
                "node_id": "document:root",
                "role": "DOCUMENT",
                "text_or_value": None,
                "parent": None,
                "location": {"document": "current"},
                "extension": {"native_epub_projection": True},
            }
        )
    accepted = {"p", "h1", "h2", "h3", "h4", "h5", "h6", "li", "td", "th"}
    tolerant_fallback_count = 0
    for section_index, name in enumerate(content_names, start=1):
        member_items: list[dict[str, object]] = []
        if query is None:
            _append_epub_projection_item(
                items,
                text=Path(name).name,
                query=None,
                kind="epub_section",
                role="SECTION",
                parent="document:root",
                location={
                    "section": section_index,
                    "section_title": Path(name).name,
                },
                budget=budget,
                node_id=f"section:{section_index}",
            )
        item_index = 0
        try:
            with archive.open(name) as source:
                for _event, element in ElementTree.iterparse(source, events=("end",)):
                    local_name = _xml_local_name(element.tag)
                    if local_name not in accepted:
                        continue
                    item_index += 1
                    _append_epub_projection_item(
                        member_items,
                        text=_xml_element_text(element),
                        query=query,
                        kind="epub_section_text",
                        role="HEADING" if local_name.startswith("h") else "PARAGRAPH",
                        parent=f"section:{section_index}",
                        location={
                            "section": section_index,
                            "section_title": Path(name).name,
                            "ordinal": item_index,
                        },
                        budget=budget,
                    )
                    element.clear()
        except ElementTree.ParseError:
            tolerant_fallback_count += 1
            member_items = _tolerant_epub_html_member(archive, name, query, section_index, budget)
            safe_title = Path(name).name[:160]
            warnings.append(f"EPUB_HTML_TOLERANT_FALLBACK:{section_index}:{safe_title}")
        items.extend(member_items)
    if tolerant_fallback_count:
        warnings.insert(0, f"EPUB_HTML_TOLERANT_FALLBACK_COUNT:{tolerant_fallback_count}")
    if budget is not None:
        warnings.append(f"EPUB_NATIVE_VIEW:{native_view}")
        if budget.omitted_items:
            warnings.append(f"EPUB_NATIVE_TEXT_ITEMS_OMITTED:{budget.omitted_items}")
    return items, len(content_names), warnings


def _streaming_epub_query_map(
    archive: zipfile.ZipFile, query: str
) -> tuple[list[dict[str, object]], int, list[str]]:
    return _epub_projection(archive, query)


@dataclass(frozen=True, slots=True)
class NativeEpubWorker:
    view: str

    def __call__(self, source_path: str) -> dict[str, Any]:
        return _native_epub_worker(source_path, self.view)


def _native_epub_worker(source_path: str, view: str = "READ") -> dict[str, Any]:
    """Project EPUB spine content with tolerant HTML fallback and no network access."""
    with zipfile.ZipFile(source_path) as archive:
        items, container_count, warnings = _epub_projection(archive, None, native_view=view)
    return {
        "backend_name": "STEWARDNativeEpub",
        "backend_version": version("local-system-steward"),
        "warnings": [f"EPUB_NATIVE_CONTAINERS:{container_count}", *warnings],
        "items": items,
    }


def _streaming_package_query_map(
    source_path: str, source_format: str, query: str
) -> dict[str, Any]:
    with zipfile.ZipFile(source_path) as archive:
        if source_format == "DOCX":
            items, container_count, warnings = _streaming_docx_query_map(archive, query)
        elif source_format == "PPTX":
            items, container_count, warnings = _streaming_pptx_query_map(archive, query)
        elif source_format == "EPUB":
            items, container_count, warnings = _streaming_epub_query_map(archive, query)
        else:
            raise RuntimeError("streaming package query map does not support this format")
    return {
        "backend_name": "STEWARDStreamingMap",
        "backend_version": version("local-system-steward"),
        "warnings": [f"STREAMING_QUERY_MAP_CONTAINERS:{container_count}", *warnings],
        "items": items,
    }


@dataclass(frozen=True, slots=True)
class StreamingQueryMapWorker:
    """Pickle-safe query scanner selected only for sufficiently large sources."""

    source_format: str
    query: str

    def __call__(self, source_path: str) -> dict[str, Any]:
        if self.source_format == "PDF":
            return _streaming_pdf_query_map(source_path, self.query)
        if self.source_format == "XLSX":
            return _streaming_xlsx_query_map(source_path, self.query)
        if self.source_format in {"DOCX", "PPTX", "EPUB"}:
            return _streaming_package_query_map(source_path, self.source_format, self.query)
        raise RuntimeError("streaming query map does not support this format")


def _pymupdf4llm_pdf_worker(source_path: str) -> dict[str, Any]:
    """Extract safe PDF layout and local OCR without retaining page images."""
    import pymupdf4llm  # type: ignore[import-untyped]

    chunks = pymupdf4llm.to_markdown(
        source_path,
        page_chunks=True,
        use_ocr=True,
        force_ocr=False,
        write_images=False,
        embed_images=False,
    )
    if not isinstance(chunks, list):
        raise RuntimeError("PDF backend returned an invalid page projection")
    items: list[dict[str, object]] = []
    for page_index, chunk in enumerate(chunks, start=1):
        if not isinstance(chunk, dict):
            raise RuntimeError("PDF backend returned an invalid page chunk")
        text = chunk.get("text", "")
        if not isinstance(text, str):
            raise RuntimeError("PDF backend returned non-text page content")
        extension: dict[str, object] = {"page": page_index, "region_kind": "page"}
        page_boxes = chunk.get("page_boxes")
        if isinstance(page_boxes, list):
            extension["page_box_count"] = len(page_boxes)
            regions = _pdf_regions(page_boxes)
            if regions:
                extension["regions"] = regions
        items.append(
            {
                "kind": "pdf_page_block",
                "text_or_value": text,
                "parent": None,
                "location": {"page": page_index, "block": 1},
                "extension": extension,
            }
        )
    native_items, native_warnings = project_pdf_native(source_path)
    return {
        "backend_name": "PyMuPDF4LLM",
        "backend_version": version("pymupdf4llm"),
        "warnings": native_warnings,
        "items": [*items, *native_items],
    }


def _pymupdf_pdf_structure_worker(source_path: str) -> dict[str, Any]:
    """Project PDF identity and native hierarchy without parsing every page body."""
    pymupdf = import_module("pymupdf")
    document = pymupdf.open(source_path)
    try:
        page_count = document.page_count
        repaired = bool(document.is_repaired)
        encrypted = bool(document.is_encrypted)
    finally:
        document.close()
    native_items, native_warnings = project_pdf_native(
        source_path,
        include_page_auxiliary=False,
    )
    items: list[dict[str, object]] = [
        {
            "kind": "pdf_document",
            "role": "DOCUMENT",
            "text_or_value": None,
            "node_id": "document:current",
            "parent": None,
            "location": {"document": "current"},
            "extension": {
                "page_count": page_count,
                "repaired": repaired,
                "encrypted": encrypted,
                "structure_projection": "NATIVE_OUTLINE_AND_METADATA",
                "page_body_parsed": False,
            },
        }
    ]
    items.extend(item for item in native_items if item.get("kind") != "pdf_document")
    outline_count = sum(item.get("kind") == "pdf_outline" for item in items)
    root_extension = items[0]["extension"]
    if isinstance(root_extension, dict):
        root_extension.update(
            {
                "native_outline_available": outline_count > 0,
                "native_outline_entry_count": outline_count,
                "structure_completeness": (
                    "NATIVE_OUTLINE_COMPLETE" if outline_count else "ROOT_METADATA_ONLY"
                ),
                "page_auxiliary_scanned": False,
                "inferred_headings_attempted": False,
            }
        )
    outline_warnings = [] if outline_count else ["PDF_NATIVE_STRUCTURE_OUTLINE_ABSENT"]
    return {
        "backend_name": "PyMuPDFNativeStructure",
        "backend_version": version("pymupdf"),
        "warnings": [
            "PDF_NATIVE_STRUCTURE_BODY_NOT_PARSED",
            f"PDF_NATIVE_STRUCTURE_OUTLINE_ENTRIES:{outline_count}",
            *outline_warnings,
            *native_warnings,
        ],
        "items": items,
    }


def _openpyxl_xlsx_worker(source_path: str) -> dict[str, Any]:
    """Extract safe workbook structure without evaluating formulas or following links."""
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    # ``read_only=False`` is necessary for table/chart identities. The worker
    # never saves or mutates this in-memory workbook; formulas remain raw text.
    workbook = load_workbook(
        source_path,
        read_only=False,
        data_only=False,
        keep_links=False,
        keep_vba=False,
    )
    items: list[dict[str, object]] = [
        {
            "kind": "xlsx_workbook",
            "text_or_value": None,
            "parent": None,
            "location": {"workbook": "current"},
            "extension": {"sheet_count": len(workbook.worksheets)},
        }
    ]
    for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
        sheet_parent = f"sheet:{sheet_index}"
        items.append(
            {
                "kind": "xlsx_sheet",
                "role": "SHEET",
                "text_or_value": sheet.title,
                "parent": "workbook:current",
                "location": {"sheet": sheet.title, "sheet_index": sheet_index},
                "extension": {
                    "sheet_order": sheet_index,
                    "table_count": len(sheet.tables),
                    "chart_count": len(sheet._charts),
                    "merged_range_count": len(sheet.merged_cells.ranges),
                    "hidden": sheet.sheet_state != "visible",
                },
            }
        )
        for merged_index, merged_range in enumerate(sheet.merged_cells.ranges, start=1):
            items.append(
                {
                    "kind": "xlsx_merged_range",
                    "role": "SECTION",
                    "text_or_value": str(merged_range),
                    "parent": sheet_parent,
                    "location": {
                        "sheet": sheet.title,
                        "sheet_index": sheet_index,
                        "range": str(merged_range),
                    },
                    "extension": {"merged_range_index": merged_index},
                }
            )
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None and cell.comment is None:
                    continue
                if cell.value is not None:
                    formula = isinstance(cell.value, str) and cell.value.startswith("=")
                    items.append(
                        {
                            "kind": "xlsx_cell",
                            "role": "FORMULA" if formula else "TABLE_CELL",
                            "text_or_value": _xlsx_cell_value(cell.value),
                            "parent": sheet_parent,
                            "location": {
                                "sheet": sheet.title,
                                "sheet_index": sheet_index,
                                "cell": cell.coordinate,
                            },
                            "extension": {
                                **(
                                    {"formula": cell.value}
                                    if formula
                                    else {"data_type": cell.data_type}
                                ),
                                "row_hidden": bool(sheet.row_dimensions[cell.row].hidden),
                                "column_hidden": bool(
                                    sheet.column_dimensions[cell.column_letter].hidden
                                ),
                            },
                        }
                    )
                if cell.comment is not None:
                    comment_text = str(cell.comment.text)
                    items.append(
                        {
                            "kind": "xlsx_comment",
                            "role": "NOTE",
                            "text_or_value": comment_text[:4_096],
                            "parent": sheet_parent,
                            "location": {
                                "sheet": sheet.title,
                                "sheet_index": sheet_index,
                                "cell": cell.coordinate,
                                "comment": 1,
                            },
                            "extension": {
                                "author": str(cell.comment.author)[:256],
                                "text_truncated": len(comment_text) > 4_096,
                            },
                        }
                    )
        for table_index, table in enumerate(sheet.tables.values(), start=1):
            items.append(
                {
                    "kind": "xlsx_table",
                    "role": "TABLE",
                    "text_or_value": table.name,
                    "parent": sheet_parent,
                    "location": {
                        "sheet": sheet.title,
                        "sheet_index": sheet_index,
                        "table": table.name,
                    },
                    "extension": {"table_index": table_index, "ref": table.ref},
                }
            )
        for chart_index, chart in enumerate(sheet._charts, start=1):
            chart_extension = {
                "chart_index": chart_index,
                **openpyxl_chart_extension(chart),
            }
            items.append(
                {
                    "kind": "xlsx_chart",
                    "role": "FIGURE",
                    "text_or_value": chart_search_text(chart_extension),
                    "parent": sheet_parent,
                    "location": {
                        "sheet": sheet.title,
                        "sheet_index": sheet_index,
                        "chart": chart_index,
                    },
                    "extension": chart_extension,
                }
            )
    workbook.close()
    return {
        "backend_name": "openpyxl",
        "backend_version": version("openpyxl"),
        "warnings": [],
        "items": items,
    }


_XLSX_STRUCTURED_REFERENCE = re.compile(r"(?:'[^']+'|[A-Za-z_\\][A-Za-z0-9_.\\]*)\[[^\r\n=]+?\]")


def _openpyxl_xlsx_formula_worker(source_path: str) -> dict[str, Any]:
    """Stream raw Excel formulas without evaluation or full workbook materialization."""
    from openpyxl import load_workbook

    workbook = load_workbook(
        source_path,
        read_only=True,
        data_only=False,
        keep_links=False,
        keep_vba=False,
    )
    items: list[dict[str, object]] = []
    formula_count = 0
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str) or not value.startswith("="):
                        continue
                    formula_count += 1
                    if len(items) >= MAX_NATIVE_FORMULA_ITEMS:
                        continue
                    references = tuple(dict.fromkeys(_XLSX_STRUCTURED_REFERENCE.findall(value)))
                    items.append(
                        {
                            "kind": "xlsx_formula",
                            "role": "FORMULA",
                            "text_or_value": value,
                            "parent": f"sheet:{sheet_index}",
                            "location": {
                                "sheet": sheet.title,
                                "sheet_index": sheet_index,
                                "cell": cell.coordinate,
                            },
                            "extension": {
                                "formula": value,
                                "formula_language": "EXCEL",
                                "formula_kind": (
                                    "STRUCTURED_REFERENCE" if references else "CELL_FORMULA"
                                ),
                                "structured_references": list(references),
                                "evaluated": False,
                            },
                        }
                    )
    finally:
        workbook.close()
    return {
        "backend_name": "openpyxl-formula-stream",
        "backend_version": version("openpyxl"),
        "warnings": [
            f"XLSX_FORMULA_TOTAL:{formula_count}",
            *(
                [f"XLSX_FORMULA_OMITTED:{formula_count - len(items)}"]
                if formula_count > len(items)
                else []
            ),
        ],
        "items": items,
    }


def _xlsx_cell_value(value: object) -> str:
    """Keep values observational and JSON-safe without formula evaluation."""
    if isinstance(value, str):
        return value
    if isinstance(value, (bool, int, float)):
        return str(value)
    return str(value)


def _python_pptx_worker(source_path: str) -> dict[str, Any]:
    """Extract deterministic presentation structure without rendering or media reads."""
    presentation_factory = getattr(import_module("pptx"), "Presentation")

    try:
        presentation = presentation_factory(source_path)
    except (AttributeError, KeyError, OSError, ValueError, zipfile.BadZipFile) as error:
        # python-pptx raises a few implementation-specific errors for a ZIP
        # that has PPTX evidence but lacks a valid package graph. Normalize the
        # recognized source, while preserving unexpected extraction failures.
        raise _RecognizedDocumentMalformedError from error
    items: list[dict[str, object]] = [
        {
            "kind": "pptx_presentation",
            "text_or_value": None,
            "parent": None,
            "location": {"presentation": "current"},
            "extension": {"slide_count": len(presentation.slides)},
        }
    ]
    for slide_index, slide in enumerate(presentation.slides, start=1):
        slide_parent = f"slide:{slide_index}"
        ordered_shapes = sorted(
            enumerate(slide.shapes, start=1),
            key=lambda pair: (
                int(pair[1].top),
                int(pair[1].left),
                pair[0],
            ),
        )
        items.append(
            {
                "kind": "pptx_slide",
                "text_or_value": None,
                "parent": "presentation:current",
                "location": {"slide": slide_index},
                "extension": {"slide_order": slide_index, "shape_count": len(slide.shapes)},
            }
        )
        for reading_order, (shape_index, shape) in enumerate(ordered_shapes, start=1):
            shape_parent = f"slide:{slide_index}:shape:{shape_index}"
            shape_type = str(shape.shape_type)
            location = {"slide": slide_index, "shape": shape_index}
            geometry = {
                "left": int(shape.left),
                "top": int(shape.top),
                "width": int(shape.width),
                "height": int(shape.height),
            }
            accessibility = pptx_shape_accessibility(shape)
            items.append(
                {
                    "kind": "pptx_shape",
                    "role": "FIGURE" if not shape.has_text_frame else "OTHER",
                    "text_or_value": None,
                    "parent": slide_parent,
                    "location": location,
                    "extension": {
                        "shape_type": shape_type,
                        "reading_order": reading_order,
                        "geometry_emu": geometry,
                        "accessibility": accessibility,
                    },
                }
            )
            if any(key in accessibility for key in ("title", "description")):
                items.append(
                    {
                        "kind": "pptx_accessibility",
                        "role": "CAPTION",
                        "text_or_value": "\n".join(accessibility.values()),
                        "parent": shape_parent,
                        "location": {**location, "accessibility": 1},
                        "extension": accessibility,
                    }
                )
            if shape.has_text_frame and shape.text:
                items.append(
                    {
                        "kind": "pptx_text",
                        "role": "PARAGRAPH",
                        "text_or_value": shape.text,
                        "parent": shape_parent,
                        "location": location,
                        "extension": {
                            "shape_type": shape_type,
                            "reading_order": reading_order,
                            "geometry_emu": geometry,
                            "accessibility": accessibility,
                        },
                    }
                )
            if shape.has_table:
                table = shape.table
                table_parent = f"{shape_parent}:table:1"
                items.append(
                    {
                        "kind": "pptx_table",
                        "role": "TABLE",
                        "text_or_value": None,
                        "parent": shape_parent,
                        "location": {**location, "table": 1},
                        "extension": {"rows": len(table.rows), "columns": len(table.columns)},
                    }
                )
                for row_index, row in enumerate(table.rows, start=1):
                    for column_index, cell in enumerate(row.cells, start=1):
                        items.append(
                            {
                                "kind": "pptx_table_cell",
                                "role": "TABLE_CELL",
                                "text_or_value": cell.text,
                                "parent": table_parent,
                                "location": {
                                    **location,
                                    "table": 1,
                                    "row": row_index,
                                    "column": column_index,
                                },
                                "extension": {
                                    "is_merge_origin": bool(cell.is_merge_origin),
                                    "is_spanned": bool(cell.is_spanned),
                                    "span_height": int(cell.span_height),
                                    "span_width": int(cell.span_width),
                                },
                            }
                        )
            if shape.has_chart:
                chart = shape.chart
                chart_extension = pptx_chart_extension(chart)
                items.append(
                    {
                        "kind": "pptx_chart",
                        "role": "FIGURE",
                        "text_or_value": chart_search_text(chart_extension),
                        "parent": shape_parent,
                        "location": {**location, "chart": 1},
                        "extension": chart_extension,
                    }
                )
        if slide.has_notes_slide:
            notes_text = slide.notes_slide.notes_text_frame.text.strip()
            if notes_text:
                items.append(
                    {
                        "kind": "pptx_speaker_notes",
                        "role": "NOTE",
                        "text_or_value": notes_text,
                        "parent": slide_parent,
                        "location": {"slide": slide_index, "notes": 1},
                        "extension": {"speaker_notes": True},
                    }
                )
    return {
        "backend_name": "python-pptx",
        "backend_version": version("python-pptx"),
        "warnings": [],
        "items": items,
    }


def _chart_series_name(series: object) -> str | None:
    """Return only safe series-label text; embedded workbook data stays unparsed."""
    name = getattr(series, "name", None)
    return name if isinstance(name, str) else None


def _markitdown_docx_worker(source_path: str) -> dict[str, Any]:
    """Project one validated DOCX through MarkItDown's DOCX converter only."""
    converter_type = getattr(
        import_module("markitdown.converters._docx_converter"), "DocxConverter"
    )
    stream_info_type = getattr(import_module("markitdown._stream_info"), "StreamInfo")
    try:
        with Path(source_path).open("rb") as source:
            projection = converter_type().convert(
                source,
                stream_info_type(
                    extension=".docx",
                    mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                ),
            )
    except (AttributeError, KeyError, OSError, ValueError, zipfile.BadZipFile) as error:
        raise _RecognizedDocumentMalformedError from error
    markdown = getattr(projection, "markdown", None)
    if not isinstance(markdown, str):
        raise RuntimeError("MarkItDown returned an invalid DOCX projection")
    items, warnings = _normalized_docx_projection(markdown)
    auxiliary_items, auxiliary_warnings = project_docx_auxiliary(source_path)
    return {
        "backend_name": "MarkItDown",
        "backend_version": version("markitdown"),
        "warnings": [*warnings, *auxiliary_warnings],
        "items": [*items, *auxiliary_items],
    }


def _normalized_docx_projection(markdown: str) -> tuple[list[dict[str, object]], list[str]]:
    """Strip binary-like projection fields and retain a compact DOCX block model."""
    items: list[dict[str, object]] = [
        {
            "kind": "docx_document",
            "text_or_value": None,
            "parent": None,
            "location": {"document": "current"},
            "extension": {"projection": "restricted"},
        }
    ]
    warnings: list[str] = []
    data_url_count = 0
    binary_line_count = 0
    lines = markdown.splitlines()
    index = 0
    block_index = 0
    heading_stack: dict[int, str] = {}

    def current_parent() -> str:
        if not heading_stack:
            return "document:current"
        return heading_stack[max(heading_stack)]

    while index < len(lines):
        line = lines[index]
        sanitized, replacements = _DOCX_DATA_URL_PATTERN.subn("", line)
        if replacements:
            data_url_count += replacements
            block_index += 1
            items.append(
                {
                    "kind": "docx_image_reference",
                    "text_or_value": None,
                    "parent": "document:current",
                    "location": {"block": block_index},
                    "extension": {"embedded_payload_stripped": True},
                }
            )
            if not sanitized.strip() or sanitized.strip().startswith("![]("):
                index += 1
                continue
            line = sanitized
        if not line.strip():
            index += 1
            continue
        if _looks_like_docx_raw_base64(line.strip()):
            binary_line_count += 1
            index += 1
            continue
        if line.strip().startswith("|") and line.strip().endswith("|"):
            table_lines: list[str] = []
            while (
                index < len(lines)
                and lines[index].strip().startswith("|")
                and lines[index].strip().endswith("|")
            ):
                table_lines.append(lines[index].strip())
                index += 1
            table_rows = _docx_table_rows(table_lines)
            if table_rows:
                block_index += 1
                table_node = f"document:block:{block_index}:table:1"
                items.append(
                    {
                        "kind": "docx_table",
                        "node_id": table_node,
                        "role": "TABLE",
                        "text_or_value": None,
                        "parent": current_parent(),
                        "location": {"block": block_index, "table": 1},
                        "extension": {
                            "rows": len(table_rows),
                            "columns": max(len(row) for row in table_rows),
                        },
                    }
                )
                for row_index, row in enumerate(table_rows, start=1):
                    for column_index, value in enumerate(row, start=1):
                        items.append(
                            {
                                "kind": "docx_table_cell",
                                "role": "TABLE_CELL",
                                "text_or_value": value,
                                "parent": table_node,
                                "location": {
                                    "block": block_index,
                                    "table": 1,
                                    "row": row_index,
                                    "column": column_index,
                                },
                                "extension": None,
                            }
                        )
            continue
        block_index += 1
        heading = _DOCX_HEADING_PATTERN.match(line.strip())
        if heading is not None:
            level = len(heading.group(1))
            parent_candidates = [candidate for candidate in heading_stack if candidate < level]
            parent = (
                heading_stack[max(parent_candidates)] if parent_candidates else "document:current"
            )
            node_id = f"document:block:{block_index}:heading:{level}"
            heading_stack = {
                candidate: value for candidate, value in heading_stack.items() if candidate < level
            }
            heading_stack[level] = node_id
            items.append(
                {
                    "kind": "docx_heading",
                    "node_id": node_id,
                    "role": "HEADING",
                    "text_or_value": heading.group(2),
                    "parent": parent,
                    "location": {"block": block_index},
                    "extension": {"level": level},
                }
            )
        else:
            items.append(
                {
                    "kind": "docx_paragraph",
                    "role": "PARAGRAPH",
                    "text_or_value": line.strip(),
                    "parent": current_parent(),
                    "location": {"block": block_index},
                    "extension": None,
                }
            )
        index += 1
    if data_url_count:
        warnings.append(f"embedded_data_urls_stripped:{data_url_count}")
    if binary_line_count:
        warnings.append(f"binary_like_projection_lines_stripped:{binary_line_count}")
    return items, warnings


def _docx_table_rows(lines: list[str]) -> list[list[str]]:
    """Convert Markdown table lines to values without retaining Markdown syntax."""
    rows: list[list[str]] = []
    for line in lines:
        values = [value.strip() for value in line.strip("|").split("|")]
        if not any(values):
            continue
        if values and all(value and set(value) <= {"-", ":"} for value in values):
            continue
        rows.append(values)
    return rows


def _looks_like_docx_raw_base64(value: str) -> bool:
    """Avoid treating ordinary long text as an embedded binary payload."""
    return (
        len(value) >= 128
        and len(value) % 4 == 0
        and _DOCX_BINARY_LINE_PATTERN.fullmatch(value) is not None
    )


def _pdf_regions(page_boxes: list[object]) -> list[dict[str, object]]:
    """Project only safe page-box facts; omit backend metadata and staging paths."""
    regions: list[dict[str, object]] = []
    for raw in page_boxes:
        if not isinstance(raw, dict):
            continue
        index = raw.get("index")
        region_class = raw.get("class")
        bbox = raw.get("bbox")
        if (
            isinstance(index, int)
            and isinstance(region_class, str)
            and isinstance(bbox, (list, tuple))
            and len(bbox) == 4
            and all(
                isinstance(value, (int, float)) and not isinstance(value, bool) for value in bbox
            )
        ):
            regions.append(
                {
                    "region_index": index,
                    "region_class": region_class,
                    "bbox": [float(value) for value in bbox],
                }
            )
    return regions


def _normalized_items(value: object) -> tuple[NormalizedDocumentItem, ...] | None:
    if not isinstance(value, list):
        return None
    items: list[NormalizedDocumentItem] = []
    for raw in value:
        if not isinstance(raw, dict):
            return None
        kind = raw.get("kind")
        text = raw.get("text_or_value")
        parent = raw.get("parent")
        location = raw.get("location")
        extension = raw.get("extension")
        node_id = raw.get("node_id")
        role = raw.get("role")
        if (
            not isinstance(kind, str)
            or (text is not None and not isinstance(text, str))
            or (parent is not None and not isinstance(parent, str))
            or not isinstance(location, dict)
            or any(
                not isinstance(key, str) or not isinstance(item, (int, str))
                for key, item in location.items()
            )
            or (
                extension is not None
                and (not isinstance(extension, dict) or not _safe_json_value(extension))
            )
            or (node_id is not None and not isinstance(node_id, str))
            or (role is not None and not isinstance(role, str))
        ):
            return None
        ordinal = len(items) + 1
        items.append(
            NormalizedDocumentItem(
                kind,
                text,
                parent,
                dict(location),
                dict(extension) if extension else None,
                node_id or f"document:item:{ordinal}",
                role or _legacy_item_role(kind),
            )
        )
    return tuple(items)


def _legacy_item_role(kind: str) -> str:
    normalized = kind.lower()
    if "heading" in normalized or normalized.endswith("_title"):
        return "HEADING"
    if "table_cell" in normalized or normalized == "xlsx_cell":
        return "TABLE_CELL"
    if "table" in normalized:
        return "TABLE"
    if "chart" in normalized or "image" in normalized or "picture" in normalized:
        return "FIGURE"
    if "sheet" in normalized:
        return "SHEET"
    if "slide" in normalized:
        return "SLIDE"
    if normalized == "pdf_page_block":
        return "SECTION"
    if normalized.endswith(("_document", "_workbook", "_presentation")):
        return "DOCUMENT"
    return "PARAGRAPH"


def _format_native_items(
    items: tuple[NormalizedDocumentItem, ...], source_format: str
) -> tuple[NormalizedDocumentItem, ...]:
    """Add observed container coordinates without replacing backend graph identity."""

    projected: list[NormalizedDocumentItem] = []
    section_index = 0
    section_title: str | None = None
    for item in items:
        location = dict(item.location)
        if source_format in IMAGE_SOURCE_FORMATS and "page" not in location:
            location["page"] = 1
        if source_format in {"EPUB", "DOCX"}:
            role = item.role or _legacy_item_role(item.kind)
            if role == "HEADING" and item.text_or_value:
                section_index += 1
                section_title = item.text_or_value[:256]
            if section_index:
                location["section"] = section_index
            if section_title is not None:
                location["section_title"] = section_title
        projected.append(replace(item, location=location))
    return tuple(projected)


def _safe_json_value(value: object) -> bool:
    if value is None or isinstance(value, (bool, int, float, str)):
        return True
    if isinstance(value, list):
        return all(_safe_json_value(item) for item in value)
    if isinstance(value, dict):
        return all(isinstance(key, str) and _safe_json_value(item) for key, item in value.items())
    return False


@dataclass(slots=True)
class StructuredDocumentParserAdapter:
    """Backend-neutral execution boundary for the accepted offline format slices."""

    ingress: ProjectOwnedBoundedDocumentIngress
    worker: IsolatedPdfWorker = None  # type: ignore[assignment]
    xlsx_worker: IsolatedXlsxWorker = None  # type: ignore[assignment]
    pptx_worker: IsolatedPptxWorker = None  # type: ignore[assignment]
    docx_worker: IsolatedDocxWorker = None  # type: ignore[assignment]
    docling_worker: IsolatedDoclingWorker = None  # type: ignore[assignment]
    enriched_worker: IsolatedEnrichedDoclingWorker = None  # type: ignore[assignment]
    macos_ocr_worker: IsolatedMacOcrWorker = None  # type: ignore[assignment]
    parse_cache: BoundedDocumentParseCache[_WorkerExecution] | None = None

    def __post_init__(self) -> None:
        if self.worker is None:
            self.worker = IsolatedPdfWorker()
        if self.xlsx_worker is None:
            self.xlsx_worker = IsolatedXlsxWorker()
        if self.pptx_worker is None:
            self.pptx_worker = IsolatedPptxWorker()
        if self.docx_worker is None:
            self.docx_worker = IsolatedDocxWorker()
        if self.docling_worker is None:
            self.docling_worker = IsolatedDoclingWorker()
        if self.enriched_worker is None:
            self.enriched_worker = IsolatedEnrichedDoclingWorker()
        if self.macos_ocr_worker is None:
            self.macos_ocr_worker = IsolatedMacOcrWorker()

    def observe(self, arguments: dict[str, object]) -> NormalizedDocumentObservation:
        self.ingress.preflight(arguments)
        admitted = self.ingress.admit(arguments)
        if isinstance(admitted, _IngressFailure):
            return self._failure(
                admitted.status,
                admitted.scope_id,
                admitted.relative_path,
                None,
                None,
                None,
                DocumentResourceUsage(admitted.source_bytes, 0, 0, None, 0, 0),
            )
        try:
            return self._observe_admitted(arguments, admitted)
        finally:
            admitted.close()

    def _observe_admitted(
        self,
        arguments: dict[str, object],
        admitted: _AdmittedDocumentSource,
    ) -> NormalizedDocumentObservation:
        identification_started = monotonic()
        identified = identify_document_format(admitted._staged_path, admitted.relative_path)
        identification_elapsed_ms = int((monotonic() - identification_started) * 1_000)
        admitted.admission_profile = identified.admission_profile
        admitted.source_limit_bytes = identified.source_limit_bytes
        admitted.expanded_limit_bytes = identified.expanded_limit_bytes
        admitted.archive_member_count = identified.archive_member_count
        admitted.identification_elapsed_ms = identification_elapsed_ms
        if identified.status is not None:
            return self._failure(
                identified.status,
                admitted.scope_id,
                admitted.relative_path,
                admitted.source_sha256,
                identified.source_format,
                None,
                DocumentResourceUsage(
                    admitted.source_bytes,
                    identified.expanded_bytes,
                    0,
                    None,
                    0,
                    0,
                    identified.admission_profile,
                    identified.source_limit_bytes,
                    identified.expanded_limit_bytes,
                    identified.archive_member_count,
                    ingress_elapsed_ms=admitted.ingress_elapsed_ms,
                    identification_elapsed_ms=identification_elapsed_ms,
                    operation_elapsed_ms=(admitted.ingress_elapsed_ms + identification_elapsed_ms),
                ),
                identified.reason,
            )
        source_format = identified.source_format
        if source_format is None:
            return self._failure(
                "UNSUPPORTED_FORMAT",
                admitted.scope_id,
                admitted.relative_path,
                admitted.source_sha256,
                None,
                None,
                DocumentResourceUsage(
                    admitted.source_bytes,
                    identified.expanded_bytes,
                    0,
                    None,
                    0,
                    0,
                    identified.admission_profile,
                    identified.source_limit_bytes,
                    identified.expanded_limit_bytes,
                    identified.archive_member_count,
                    ingress_elapsed_ms=admitted.ingress_elapsed_ms,
                    identification_elapsed_ms=identification_elapsed_ms,
                    operation_elapsed_ms=(admitted.ingress_elapsed_ms + identification_elapsed_ms),
                ),
                "UNACCEPTED_FORMAT",
            )
        profile = arguments.get("parser_profile", "FAST")
        if profile not in {"FAST", "AUTO", "DEEP", "ENRICHED"}:
            raise RuntimeFailure("INVALID_ARGUMENT", "parser profile is invalid")
        view = arguments.get("view", "READ")
        if view not in {"READ", "STRUCTURE", "TABLES", "FORMULAS"}:
            raise RuntimeFailure("INVALID_ARGUMENT", "document view is invalid")
        intent = arguments.get("intent", view)
        if intent not in {
            "READ",
            "STRUCTURE",
            "LOCATE",
            "EVIDENCE",
            "TABLES",
            "FORMULAS",
        }:
            raise RuntimeFailure("INVALID_ARGUMENT", "document intent is invalid")
        parser_timeout_seconds = arguments.get("parser_timeout_seconds")
        if parser_timeout_seconds is not None and (
            isinstance(parser_timeout_seconds, bool)
            or not isinstance(parser_timeout_seconds, (int, float))
            or not 1 <= parser_timeout_seconds <= 600
        ):
            raise RuntimeFailure("INVALID_ARGUMENT", "parser timeout is invalid")
        parser_deadline = (
            monotonic() + float(parser_timeout_seconds)
            if parser_timeout_seconds is not None
            else None
        )

        def remaining_parser_timeout() -> float | None:
            if parser_deadline is None:
                return None
            return max(0.001, parser_deadline - monotonic())

        if source_format in AUDIO_SOURCE_FORMATS:
            return self._observe_audio_admitted(
                arguments,
                admitted,
                source_format,
                profile=profile,
                view=view,
                intent=intent,
                parser_timeout_seconds=remaining_parser_timeout(),
            )
        if source_format in VIDEO_SOURCE_FORMATS:
            return self._observe_video_admitted(
                arguments,
                admitted,
                source_format,
                profile=profile,
                view=view,
                intent=intent,
                parser_timeout_seconds=remaining_parser_timeout(),
            )
        suffix_by_format = {
            "PDF": ".pdf",
            "EPUB": ".epub",
            "DOCX": ".docx",
            "XLSX": ".xlsx",
            "PPTX": ".pptx",
            "PNG": ".png",
            "JPEG": ".jpg",
            "TIFF": ".tiff",
        }
        suffix = suffix_by_format[source_format]
        query = arguments.get("content_query")
        pdf_native_text_probe = (
            _pdf_native_text_probe(admitted._staged_path)
            if source_format == "PDF"
            and profile == "AUTO"
            and intent in {"READ", "LOCATE", "EVIDENCE"}
            else None
        )
        queried_pdf_page_count = (
            pdf_native_text_probe[0] if pdf_native_text_probe is not None else None
        )
        queried_pdf_low_text_pages = (
            pdf_native_text_probe[1] if pdf_native_text_probe is not None else 0
        )
        queried_pdf_map = (
            source_format == "PDF"
            and intent in {"LOCATE", "EVIDENCE"}
            and (
                admitted.source_bytes >= STREAMING_QUERY_MAP_THRESHOLD_BYTES
                or (queried_pdf_page_count or 0) >= STREAMING_QUERY_MAP_PDF_PAGE_THRESHOLD
                or queried_pdf_low_text_pages > 0
            )
        )
        scanned_pdf_read = (
            source_format == "PDF"
            and intent == "READ"
            and queried_pdf_page_count is not None
            and queried_pdf_page_count > 0
            and queried_pdf_low_text_pages * 2 >= queried_pdf_page_count
        )
        threshold_query_map = (
            intent in {"LOCATE", "EVIDENCE"}
            and source_format in {"DOCX", "XLSX", "PPTX"}
            and admitted.source_bytes >= STREAMING_QUERY_MAP_THRESHOLD_BYTES
        )
        epub_query_map = source_format == "EPUB" and intent in {"LOCATE", "EVIDENCE"}
        streaming_query_map = (
            profile == "AUTO"
            and isinstance(query, str)
            and (queried_pdf_map or threshold_query_map or epub_query_map)
        )
        initial_profile = (
            "OCR_NATIVE"
            if scanned_pdf_read
            else (
                "MAP"
                if streaming_query_map
                else initial_document_profile(source_format, profile, view, intent)
            )
        )
        quality_view = "READ" if intent in {"LOCATE", "EVIDENCE"} else view
        result, cache_status = self._parse_profile(
            admitted,
            source_format,
            identified.expanded_bytes,
            suffix,
            initial_profile,
            query=query if streaming_query_map and isinstance(query, str) else None,
            parser_timeout_seconds=remaining_parser_timeout(),
        )
        attempts = [
            self._attempt(initial_profile, cache_status, result, source_format, quality_view)
        ]
        selected = result
        selected_profile: str | None = initial_profile if result.status == "COMPLETE" else None
        escalation_reason: str | None = None
        execution_selection: DocumentExecutionSelection | None = None

        if profile == "AUTO":
            if (
                intent == "EVIDENCE"
                and initial_profile in {"FAST", "MAP"}
                and result.status == "COMPLETE"
                and isinstance(query, str)
            ):
                evidence_page = arguments.get("evidence_page")
                execution_selection = self._targeted_evidence_selection(
                    result.items,
                    query,
                    evidence_page if isinstance(evidence_page, int) else None,
                    source_format=source_format,
                    map_profile=initial_profile,
                )
                native_fidelity_match = self._items_contain_native_fidelity_query(
                    result.items, query
                )
                if (
                    source_format == "PDF"
                    and execution_selection.selected_page_start is not None
                    and not native_fidelity_match
                    and initial_profile != "MAP"
                ):
                    escalation_reason = "EVIDENCE_TARGETED_PAGE_RANGE"
                    page_range = (
                        execution_selection.selected_page_start,
                        execution_selection.selected_page_end
                        or execution_selection.selected_page_start,
                    )
                    deep, deep_cache = self._parse_profile(
                        admitted,
                        source_format,
                        identified.expanded_bytes,
                        suffix,
                        "DEEP",
                        page_range=page_range,
                        parser_timeout_seconds=remaining_parser_timeout(),
                    )
                    attempts.append(
                        self._attempt("DEEP", deep_cache, deep, source_format, quality_view)
                    )
                    if deep.status == "COMPLETE" and self._items_contain_query(deep.items, query):
                        selected = self._add_warning(deep, "EVIDENCE_TARGETED_DEEP_PARSE")
                        selected_profile = "DEEP"
                    elif deep.status == "COMPLETE":
                        selected = self._add_warning(
                            result, "EVIDENCE_TARGETED_DEEP_QUERY_NOT_RETAINED"
                        )
                elif (
                    source_format != "PDF"
                    and initial_profile != "MAP"
                    and not native_fidelity_match
                    and (
                        not execution_selection.matched_container_ids
                        or any(
                            item.quality.status == "INSUFFICIENT"
                            for item in execution_selection.container_qualities
                        )
                    )
                ):
                    escalation_reason = "EVIDENCE_QUERY_CONTAINER_INSUFFICIENT"
                    deep, deep_cache = self._parse_profile(
                        admitted,
                        source_format,
                        identified.expanded_bytes,
                        suffix,
                        "DEEP",
                        parser_timeout_seconds=remaining_parser_timeout(),
                    )
                    attempts.append(
                        self._attempt("DEEP", deep_cache, deep, source_format, quality_view)
                    )
                    if deep.status == "COMPLETE" and self._items_contain_query(deep.items, query):
                        selected = self._add_warning(deep, "EVIDENCE_QUERY_RELEVANT_DEEP_FALLBACK")
                        selected_profile = "DEEP"
                        execution_selection = replace(
                            self._targeted_evidence_selection(
                                deep.items,
                                query,
                                source_format=source_format,
                            ),
                            strategy=("FAST_QUERY_MAP_THEN_DEEP_NATIVE_CONTAINER_QUALITY"),
                            map_profile="DEEP",
                        )
            if (
                intent == "EVIDENCE"
                and isinstance(query, str)
                and result.status == "COMPLETE"
                and source_format in {"PDF", *IMAGE_SOURCE_FORMATS}
                and not self._items_contain_native_fidelity_query(result.items, query)
            ):
                evidence_page = arguments.get("evidence_page")
                ocr_page_range: tuple[int, int] | None = None
                ocr_reason: str | None = None
                # MAP already performs page-local OCR for low-text PDF pages while
                # scanning for the query.  Re-running whole-document OCR after a
                # MAP miss would duplicate the expensive work without a new target.
                if attempts[0].quality.status == "INSUFFICIENT" and initial_profile != "MAP":
                    ocr_reason = "EVIDENCE_NATIVE_TEXT_INSUFFICIENT"
                    if source_format == "PDF" and isinstance(evidence_page, int):
                        ocr_page_range = (evidence_page, evidence_page)
                elif (
                    source_format == "PDF"
                    and execution_selection is not None
                    and execution_selection.selected_page_start is not None
                    and attempts[-1].profile == "DEEP"
                    and attempts[-1].quality.status == "INSUFFICIENT"
                ):
                    ocr_reason = "EVIDENCE_TARGETED_PAGE_TEXT_INSUFFICIENT"
                    ocr_page_range = (
                        execution_selection.selected_page_start,
                        execution_selection.selected_page_end
                        or execution_selection.selected_page_start,
                    )
                elif (
                    source_format in IMAGE_SOURCE_FORMATS
                    and attempts[-1].quality.status == "INSUFFICIENT"
                ):
                    ocr_reason = "EVIDENCE_IMAGE_TEXT_INSUFFICIENT"
                if ocr_reason is not None:
                    escalation_reason = ocr_reason
                    ocr, ocr_cache = self._parse_profile(
                        admitted,
                        source_format,
                        identified.expanded_bytes,
                        suffix,
                        "OCR",
                        page_range=ocr_page_range,
                        parser_timeout_seconds=remaining_parser_timeout(),
                    )
                    attempts.append(
                        self._attempt("OCR", ocr_cache, ocr, source_format, quality_view)
                    )
                    if ocr.status == "COMPLETE" and self._items_contain_query(ocr.items, query):
                        selected = self._add_warning(ocr, "EVIDENCE_QUALITY_GATED_LOCAL_OCR")
                        selected_profile = "OCR"
                        execution_selection = replace(
                            self._targeted_evidence_selection(
                                ocr.items,
                                query,
                                source_format=source_format,
                            ),
                            strategy="QUALITY_GATED_LOCAL_OCR",
                            map_profile="OCR",
                        )
            quality = attempts[0].quality
            if (
                intent != "EVIDENCE"
                and initial_profile == "FAST"
                and quality.status == "INSUFFICIENT"
            ):
                escalation_reason = quality.reason_codes[0]
                deep, deep_cache = self._parse_profile(
                    admitted,
                    source_format,
                    identified.expanded_bytes,
                    suffix,
                    "DEEP",
                    parser_timeout_seconds=remaining_parser_timeout(),
                )
                attempts.append(
                    self._attempt("DEEP", deep_cache, deep, source_format, quality_view)
                )
                if deep.status == "COMPLETE":
                    selected = self._add_warning(deep, "AUTO_ESCALATED_TO_DEEP")
                    selected_profile = "DEEP"
            if (
                intent != "EVIDENCE"
                and selected_profile == "DEEP"
                and attempts[-1].quality.status == "INSUFFICIENT"
                and source_format in {"PDF", *IMAGE_SOURCE_FORMATS}
            ):
                escalation_reason = attempts[-1].quality.reason_codes[0]
                ocr, ocr_cache = self._parse_profile(
                    admitted,
                    source_format,
                    identified.expanded_bytes,
                    suffix,
                    "OCR",
                    parser_timeout_seconds=remaining_parser_timeout(),
                )
                attempts.append(self._attempt("OCR", ocr_cache, ocr, source_format, quality_view))
                if ocr.status == "COMPLETE" and (
                    attempts[-1].quality.status == "SUFFICIENT"
                    or attempts[-1].quality.alphanumeric_characters
                    > attempts[-2].quality.alphanumeric_characters
                ):
                    selected = self._add_warning(ocr, "AUTO_ESCALATED_TO_MACOS_VISION")
                    selected_profile = "OCR"
            if (
                selected.status != "COMPLETE"
                and source_format
                in {
                    "PDF",
                    "EPUB",
                    "DOCX",
                    "XLSX",
                    "PPTX",
                }
                and initial_profile not in {"FAST", "MAP", "OCR_NATIVE", "STRUCTURE_NATIVE"}
            ):
                escalation_reason = "PRIMARY_PARSER_UNAVAILABLE"
                fast, fast_cache = self._parse_profile(
                    admitted,
                    source_format,
                    identified.expanded_bytes,
                    suffix,
                    "FAST",
                    native_view=view,
                    parser_timeout_seconds=remaining_parser_timeout(),
                )
                attempts.append(
                    self._attempt("FAST", fast_cache, fast, source_format, quality_view)
                )
                if fast.status == "COMPLETE":
                    selected = self._add_warning(fast, "AUTO_PRIMARY_UNAVAILABLE_FAST_FALLBACK")
                    selected_profile = "FAST"
        elif (
            profile == "DEEP"
            and result.status in {"PARSER_FAILED", "UNAVAILABLE"}
            and (source_format in {"PDF", "EPUB", "DOCX", "XLSX", "PPTX"})
        ):
            escalation_reason = "DEEP_PARSER_UNAVAILABLE"
            fast, fast_cache = self._parse_profile(
                admitted,
                source_format,
                identified.expanded_bytes,
                suffix,
                "FAST",
                native_view=view,
                parser_timeout_seconds=remaining_parser_timeout(),
            )
            attempts.append(self._attempt("FAST", fast_cache, fast, source_format, quality_view))
            selected = self._add_warning(fast, "DEEP_PARSER_UNAVAILABLE_FAST_FALLBACK")
            selected_profile = "FAST" if fast.status == "COMPLETE" else None
        elif profile == "ENRICHED" and result.status != "COMPLETE":
            escalation_reason = "ENRICHMENT_UNAVAILABLE"
            deep, deep_cache = self._parse_profile(
                admitted,
                source_format,
                identified.expanded_bytes,
                suffix,
                "DEEP",
                parser_timeout_seconds=remaining_parser_timeout(),
            )
            attempts.append(self._attempt("DEEP", deep_cache, deep, source_format, quality_view))
            selected = self._add_warning(deep, "ENRICHMENT_UNAVAILABLE_DEEP_FALLBACK")
            selected_profile = "DEEP" if deep.status == "COMPLETE" else None

        if identified.external_relationship_count:
            selected = self._add_warning(
                selected,
                f"external_relationships_ignored:{identified.external_relationship_count}",
            )
        return replace(
            selected,
            execution=DocumentExecutionTrace(
                profile,
                intent,
                view,
                initial_profile,
                selected_profile,
                escalation_reason,
                tuple(attempts),
                execution_selection,
            ),
        )

    def _observe_audio_admitted(
        self,
        arguments: dict[str, object],
        admitted: _AdmittedDocumentSource,
        source_format: str,
        *,
        profile: object,
        view: object,
        intent: object,
        parser_timeout_seconds: float | None,
    ) -> NormalizedDocumentObservation:
        """Run probe-only structure or one bounded VAD/ASR window."""
        if intent in {"TABLES", "FORMULAS"} or view in {"TABLES", "FORMULAS"}:
            return self._failure(
                "UNSUPPORTED_FORMAT",
                admitted.scope_id,
                admitted.relative_path,
                admitted.source_sha256,
                source_format,
                None,
                DocumentResourceUsage(
                    admitted.source_bytes,
                    0,
                    0,
                    None,
                    0,
                    0,
                    admitted.admission_profile,
                    admitted.source_limit_bytes,
                    None,
                    ingress_elapsed_ms=admitted.ingress_elapsed_ms,
                    identification_elapsed_ms=admitted.identification_elapsed_ms,
                ),
                "AUDIO_INTENT_UNSUPPORTED",
            )
        mode = "PROBE" if intent == "STRUCTURE" else "TRANSCRIBE"
        capabilities = audio_runtime_capabilities()
        ready = (
            capabilities["probe_ready"]
            if mode == "PROBE"
            else (
                capabilities["probe_ready"]
                and capabilities["vad_ready"]
                and capabilities["asr_ready"]
            )
        )
        if not ready:
            return self._failure(
                "UNAVAILABLE",
                admitted.scope_id,
                admitted.relative_path,
                admitted.source_sha256,
                source_format,
                "FFprobe" if mode == "PROBE" else "FasterWhisper",
                DocumentResourceUsage(
                    admitted.source_bytes,
                    0,
                    0,
                    None,
                    0,
                    0,
                    admitted.admission_profile,
                    admitted.source_limit_bytes,
                    None,
                    ingress_elapsed_ms=admitted.ingress_elapsed_ms,
                    identification_elapsed_ms=admitted.identification_elapsed_ms,
                    media={"runtime": capabilities},
                ),
                "AUDIO_RUNTIME_UNAVAILABLE",
            )
        language = arguments.get("audio_language")
        if language is not None and (
            not isinstance(language, str)
            or not language
            or len(language) > 16
            or not language.replace("-", "").isalpha()
        ):
            raise RuntimeFailure("INVALID_ARGUMENT", "audio language hint is invalid")
        analysis = arguments.get("audio_analysis", "TRANSCRIPT")
        if analysis not in {
            "TRANSCRIPT",
            "ALIGNED_WORDS",
            "SPEAKER_TURNS",
            "ALIGNED_WORDS_AND_SPEAKERS",
        }:
            raise RuntimeFailure("INVALID_ARGUMENT", "audio analysis mode is invalid")
        model_identity_sha256: str | None = None
        if mode != "PROBE":
            try:
                _model_path, _model_revision, model_identity_sha256 = resolve_local_audio_model()
            except AudioRuntimeUnavailable:
                model_identity_sha256 = None
        alignment_model_identity_sha256: str | None = None
        if analysis in {"ALIGNED_WORDS", "ALIGNED_WORDS_AND_SPEAKERS"}:
            if isinstance(language, str):
                try:
                    alignment_model_identity_sha256 = resolve_local_alignment_model(
                        language
                    ).identity_sha256
                except AudioAlignmentUnavailable:
                    alignment_model_identity_sha256 = None
            else:
                alignment_model_identity_sha256 = sha256(
                    canonical_json(alignment_runtime_capabilities())
                ).hexdigest()
        diarization_model_identity_sha256: str | None = None
        if analysis in {"SPEAKER_TURNS", "ALIGNED_WORDS_AND_SPEAKERS"}:
            try:
                diarization_model_identity_sha256 = (
                    resolve_local_diarization_model().identity_sha256
                )
            except AudioDiarizationUnavailable:
                diarization_model_identity_sha256 = None
        content_query_value = arguments.get("content_query")
        request_digest = audio_request_digest(
            source_sha256=admitted.source_sha256,
            scope_id=admitted.scope_id,
            relative_path=admitted.relative_path,
            intent=str(intent),
            content_query=content_query_value if isinstance(content_query_value, str) else None,
            language=language if isinstance(language, str) else None,
            model_identity_sha256=model_identity_sha256,
            analysis=str(analysis),
            alignment_model_identity_sha256=alignment_model_identity_sha256,
            diarization_model_identity_sha256=diarization_model_identity_sha256,
        )
        start_ms = 0
        continuation = arguments.get("audio_continuation")
        result_page_continuation = False
        if continuation is not None:
            if (
                not isinstance(continuation, dict)
                or continuation.get("request_digest") != request_digest
                or continuation.get("source_sha256") != admitted.source_sha256
            ):
                raise RuntimeFailure("INVALID_ARGUMENT", "audio continuation is invalid")
            if continuation.get("schema_version") == 2:
                window_start_ms = continuation.get("window_start_ms")
                if (
                    continuation.get("kind") != "RESULT_PAGE"
                    or type(window_start_ms) is not int
                    or window_start_ms < 0
                ):
                    raise RuntimeFailure("INVALID_ARGUMENT", "audio continuation is invalid")
                start_ms = window_start_ms
                result_page_continuation = True
            else:
                next_start_ms = continuation.get("next_start_ms")
                if type(next_start_ms) is not int or next_start_ms < 0:
                    raise RuntimeFailure("INVALID_ARGUMENT", "audio continuation is invalid")
                start_ms = next_start_ms
        worker = IsolatedParserWorker(
            worker_target=AudioDocumentWorker(
                source_format,
                mode,
                start_ms,
                language if isinstance(language, str) else None,
                request_digest,
                admitted.source_sha256,
                str(analysis),
            ),
            timeout_seconds=min(600.0, parser_timeout_seconds or 600.0),
            memory_bytes=6 * 1024 * 1024 * 1024,
        )
        cache_key = (
            admitted.source_sha256,
            source_format,
            f"AUDIO:{request_digest}:START:{start_ms}",
        )

        def execute() -> _WorkerExecution:
            with admitted.staged_copy(AUDIO_SUFFIX_BY_FORMAT[source_format]) as staged_path:
                return worker.run(staged_path)

        if result_page_continuation:
            worker_result = (
                self.parse_cache.get_existing(cache_key) if self.parse_cache is not None else None
            )
            if worker_result is None:
                raise RuntimeFailure(
                    "AUDIO_CONTINUATION_EXPIRED",
                    "audio result-page continuation is no longer cached",
                )
            cache_status = "HIT"
        elif self.parse_cache is None:
            worker_result = execute()
            cache_status = "DISABLED"
        else:
            worker_result, cache_status = self.parse_cache.get_or_compute(
                cache_key,
                execute,
                size_of=lambda value: len(canonical_json(value.payload or {})),
                cacheable=lambda value: value.status == "COMPLETE" and value.payload is not None,
            )
        resources = DocumentResourceUsage(
            admitted.source_bytes,
            0,
            worker_result.elapsed_ms,
            worker_result.peak_memory_bytes,
            0,
            0,
            admitted.admission_profile,
            admitted.source_limit_bytes,
            None,
            None,
            int(worker.timeout_seconds * 1_000),
            worker.memory_bytes,
            admitted.ingress_elapsed_ms,
            admitted.identification_elapsed_ms,
            admitted.ingress_elapsed_ms
            + admitted.identification_elapsed_ms
            + worker_result.elapsed_ms,
            "PARSER" if worker_result.status == "TIMEOUT" else None,
        )
        if worker_result.status != "COMPLETE":
            return self._failure(
                worker_result.status,
                admitted.scope_id,
                admitted.relative_path,
                admitted.source_sha256,
                source_format,
                "FFprobe" if mode == "PROBE" else "FasterWhisper",
                resources,
                "AUDIO_RUNTIME_UNAVAILABLE" if worker_result.status == "UNAVAILABLE" else None,
            )
        completed = self._complete(
            admitted,
            worker_result,
            resources,
            source_format,
            "FFprobe" if mode == "PROBE" else "FasterWhisper",
        )
        return replace(
            completed,
            execution=DocumentExecutionTrace(
                str(profile),
                str(intent),
                str(view),
                "AUDIO_PROBE" if mode == "PROBE" else "AUDIO_VAD_ASR",
                "AUDIO_PROBE" if mode == "PROBE" else "AUDIO_VAD_ASR",
                None,
                (
                    self._attempt(
                        "AUDIO_PROBE" if mode == "PROBE" else "AUDIO_VAD_ASR",
                        cache_status,
                        completed,
                        source_format,
                        str(view),
                    ),
                ),
            ),
        )

    def _observe_video_admitted(
        self,
        arguments: dict[str, object],
        admitted: _AdmittedDocumentSource,
        source_format: str,
        *,
        profile: object,
        view: object,
        intent: object,
        parser_timeout_seconds: float | None,
    ) -> NormalizedDocumentObservation:
        """Run the probe-only NEXT-024B video structure path."""
        if intent not in {"STRUCTURE", "READ", "LOCATE", "EVIDENCE"} or view in {
            "TABLES",
            "FORMULAS",
        }:
            return self._failure(
                "UNSUPPORTED_FORMAT",
                admitted.scope_id,
                admitted.relative_path,
                admitted.source_sha256,
                source_format,
                None,
                DocumentResourceUsage(
                    admitted.source_bytes,
                    0,
                    0,
                    None,
                    0,
                    0,
                    admitted.admission_profile,
                    admitted.source_limit_bytes,
                    None,
                    ingress_elapsed_ms=admitted.ingress_elapsed_ms,
                    identification_elapsed_ms=admitted.identification_elapsed_ms,
                ),
                "VIDEO_INTENT_UNSUPPORTED",
            )
        mode = "PROBE" if intent == "STRUCTURE" else "SCENES"
        analysis = arguments.get("video_analysis", "MULTIMODAL")
        content_query = arguments.get("content_query")
        if analysis not in {
            "SCENES",
            "SCENES_AND_OCR",
            "MULTIMODAL",
            "MULTIMODAL_AND_OCR",
        }:
            raise RuntimeFailure("INVALID_ARGUMENT", "video analysis mode is invalid")
        audio_language = arguments.get("audio_language")
        if audio_language is not None and (
            not isinstance(audio_language, str)
            or not audio_language
            or len(audio_language) > 16
            or not audio_language.replace("-", "").isalpha()
        ):
            raise RuntimeFailure("INVALID_ARGUMENT", "audio language hint is invalid")
        capabilities = video_runtime_capabilities()
        ready = capabilities["probe_ready"] and (mode == "PROBE" or capabilities["decode_ready"])
        if not ready:
            return self._failure(
                "UNAVAILABLE",
                admitted.scope_id,
                admitted.relative_path,
                admitted.source_sha256,
                source_format,
                "FFprobe",
                DocumentResourceUsage(
                    admitted.source_bytes,
                    0,
                    0,
                    None,
                    0,
                    0,
                    admitted.admission_profile,
                    admitted.source_limit_bytes,
                    None,
                    ingress_elapsed_ms=admitted.ingress_elapsed_ms,
                    identification_elapsed_ms=admitted.identification_elapsed_ms,
                    media={"runtime": capabilities},
                ),
                "VIDEO_RUNTIME_UNAVAILABLE",
            )
        expected_backend = (
            "FFprobe"
            if mode == "PROBE"
            else (
                "FFmpegMultimodal" if analysis in {"MULTIMODAL", "MULTIMODAL_AND_OCR"} else "FFmpeg"
            )
        )
        if mode == "PROBE":
            target: WorkerTarget = VideoProbeWorker(source_format)
            cache_key: tuple[str, str, str] | None = None
            result_page_continuation = False
        else:
            audio_model_identity_sha256: str | None = None
            if analysis in {"MULTIMODAL", "MULTIMODAL_AND_OCR"}:
                try:
                    (
                        _audio_model_path,
                        _audio_model_revision,
                        audio_model_identity_sha256,
                    ) = resolve_local_audio_model()
                except AudioRuntimeUnavailable:
                    audio_model_identity_sha256 = None
            request_digest = video_request_digest(
                source_sha256=admitted.source_sha256,
                scope_id=admitted.scope_id,
                relative_path=admitted.relative_path,
                intent=str(intent),
                content_query=content_query if isinstance(content_query, str) else None,
                analysis=str(analysis),
                audio_language=audio_language if isinstance(audio_language, str) else None,
                audio_model_identity_sha256=audio_model_identity_sha256,
            )
            start_ms = 0
            continuation = arguments.get("video_continuation")
            result_page_continuation = False
            if continuation is not None:
                if (
                    not isinstance(continuation, dict)
                    or continuation.get("request_digest") != request_digest
                    or continuation.get("source_sha256") != admitted.source_sha256
                ):
                    raise RuntimeFailure("INVALID_ARGUMENT", "video continuation is invalid")
                if (
                    continuation.get("schema_version")
                    == VIDEO_RESULT_PAGE_CONTINUATION_SCHEMA_VERSION
                ):
                    window_start_ms = continuation.get("window_start_ms")
                    if (
                        continuation.get("kind") != "RESULT_PAGE"
                        or type(window_start_ms) is not int
                        or window_start_ms < 0
                    ):
                        raise RuntimeFailure("INVALID_ARGUMENT", "video continuation is invalid")
                    start_ms = window_start_ms
                    result_page_continuation = True
                else:
                    next_start_ms = continuation.get("next_start_ms")
                    if type(next_start_ms) is not int or next_start_ms < 0:
                        raise RuntimeFailure("INVALID_ARGUMENT", "video continuation is invalid")
                    start_ms = next_start_ms
            include_ocr = analysis in {"SCENES_AND_OCR", "MULTIMODAL_AND_OCR"}
            if analysis in {"MULTIMODAL", "MULTIMODAL_AND_OCR"}:
                target = VideoTimelineWorker(
                    source_format,
                    start_ms,
                    include_ocr,
                    request_digest,
                    admitted.source_sha256,
                    content_query if isinstance(content_query, str) else None,
                    audio_language if isinstance(audio_language, str) else None,
                )
            else:
                target = VideoSceneWorker(
                    source_format,
                    start_ms,
                    include_ocr,
                    request_digest,
                    admitted.source_sha256,
                )
            cache_key = (
                admitted.source_sha256,
                source_format,
                f"VIDEO:{request_digest}:START:{start_ms}",
            )
        worker = IsolatedParserWorker(
            worker_target=target,
            timeout_seconds=min(
                600.0 if mode == "SCENES" else 60.0,
                parser_timeout_seconds or (600.0 if mode == "SCENES" else 60.0),
            ),
            memory_bytes=(
                6 * 1024 * 1024 * 1024
                if analysis in {"SCENES_AND_OCR", "MULTIMODAL_AND_OCR"}
                else 4 * 1024 * 1024 * 1024
                if analysis == "MULTIMODAL" and isinstance(content_query, str)
                else 2 * 1024 * 1024 * 1024
            ),
        )

        def execute() -> _WorkerExecution:
            with admitted.staged_copy(VIDEO_SUFFIX_BY_FORMAT[source_format]) as staged_path:
                return worker.run(staged_path)

        if result_page_continuation:
            assert cache_key is not None
            worker_result = (
                self.parse_cache.get_existing(cache_key) if self.parse_cache is not None else None
            )
            if worker_result is None:
                raise RuntimeFailure(
                    "VIDEO_CONTINUATION_EXPIRED",
                    "video result-page continuation is no longer cached",
                )
            cache_status = "HIT"
        elif mode == "PROBE" or self.parse_cache is None:
            worker_result = execute()
            cache_status = "DISABLED"
        else:
            assert cache_key is not None
            worker_result, cache_status = self.parse_cache.get_or_compute(
                cache_key,
                execute,
                size_of=lambda value: len(canonical_json(value.payload or {})),
                cacheable=lambda value: value.status == "COMPLETE" and value.payload is not None,
            )
        resources = DocumentResourceUsage(
            admitted.source_bytes,
            0,
            0 if cache_status == "HIT" else worker_result.elapsed_ms,
            None if cache_status == "HIT" else worker_result.peak_memory_bytes,
            0,
            0,
            admitted.admission_profile,
            admitted.source_limit_bytes,
            None,
            None,
            int(worker.timeout_seconds * 1_000),
            worker.memory_bytes,
            admitted.ingress_elapsed_ms,
            admitted.identification_elapsed_ms,
            admitted.ingress_elapsed_ms
            + admitted.identification_elapsed_ms
            + (0 if cache_status == "HIT" else worker_result.elapsed_ms),
            "PARSER" if worker_result.status == "TIMEOUT" else None,
        )
        if worker_result.status != "COMPLETE":
            return self._failure(
                worker_result.status,
                admitted.scope_id,
                admitted.relative_path,
                admitted.source_sha256,
                source_format,
                expected_backend,
                resources,
                "VIDEO_RUNTIME_UNAVAILABLE" if worker_result.status == "UNAVAILABLE" else None,
            )
        completed = self._complete(
            admitted, worker_result, resources, source_format, expected_backend
        )
        execution_profile = (
            "VIDEO_PROBE"
            if mode == "PROBE"
            else {
                "SCENES": "VIDEO_SCENES",
                "SCENES_AND_OCR": "VIDEO_SCENES_OCR",
                "MULTIMODAL": "VIDEO_MULTIMODAL",
                "MULTIMODAL_AND_OCR": "VIDEO_MULTIMODAL_OCR",
            }[str(analysis)]
        )
        return replace(
            completed,
            execution=DocumentExecutionTrace(
                str(profile),
                str(intent),
                str(view),
                execution_profile,
                execution_profile,
                None,
                (
                    self._attempt(
                        execution_profile, cache_status, completed, source_format, str(view)
                    ),
                ),
            ),
        )

    def _parse_profile(
        self,
        admitted: _AdmittedDocumentSource,
        source_format: str,
        expanded_bytes: int,
        suffix: str,
        profile: str,
        *,
        page_range: tuple[int, int] | None = None,
        query: str | None = None,
        native_view: str | None = None,
        parser_timeout_seconds: float | None = None,
    ) -> tuple[NormalizedDocumentObservation, str]:
        selected = self._worker_for(
            source_format,
            profile,
            page_range=page_range,
            query=query,
            native_view=native_view,
        )
        if selected is None:
            return (
                self._failure(
                    "UNAVAILABLE",
                    admitted.scope_id,
                    admitted.relative_path,
                    admitted.source_sha256,
                    source_format,
                    None,
                    DocumentResourceUsage(admitted.source_bytes, expanded_bytes, 0, None, 0, 0),
                    "RUNTIME_DEPENDENCY_UNAVAILABLE",
                ),
                "DISABLED" if self.parse_cache is None else "MISS",
            )
        backend_name, worker = selected
        effective_worker = (
            _adaptive_worker_budget(
                worker,
                source_bytes=admitted.source_bytes,
                expanded_bytes=expanded_bytes,
            )
            if isinstance(worker, IsolatedParserWorker)
            else worker
        )
        if (
            isinstance(effective_worker, IsolatedParserWorker)
            and parser_timeout_seconds is not None
        ):
            effective_worker = replace(
                effective_worker,
                timeout_seconds=min(effective_worker.timeout_seconds, parser_timeout_seconds),
            )

        def execute() -> _WorkerExecution:
            with admitted.staged_copy(suffix) as staged_path:
                return effective_worker.run(staged_path)

        if self.parse_cache is None:
            worker_result = execute()
            cache_status = "DISABLED"
        else:
            cache_profile = (
                profile
                if page_range is None
                else f"{profile}:PAGES:{page_range[0]}-{page_range[1]}"
            )
            if native_view is not None:
                cache_profile = f"{cache_profile}:VIEW:{native_view}"
            if query is not None:
                cache_profile = f"{cache_profile}:QUERY:{sha256(query.encode('utf-8')).hexdigest()}"
            worker_result, cache_status = self.parse_cache.get_or_compute(
                (admitted.source_sha256, source_format, cache_profile),
                execute,
                size_of=lambda value: len(canonical_json(value.payload or {})),
                cacheable=lambda value: value.status == "COMPLETE" and value.payload is not None,
            )
        resources = DocumentResourceUsage(
            admitted.source_bytes,
            expanded_bytes,
            0 if cache_status == "HIT" else worker_result.elapsed_ms,
            None if cache_status == "HIT" else worker_result.peak_memory_bytes,
            0,
            0,
            admitted.admission_profile,
            admitted.source_limit_bytes,
            admitted.expanded_limit_bytes,
            admitted.archive_member_count,
            (
                int(effective_worker.timeout_seconds * 1_000)
                if isinstance(effective_worker, IsolatedParserWorker)
                else None
            ),
            (
                effective_worker.memory_bytes
                if isinstance(effective_worker, IsolatedParserWorker)
                else None
            ),
            ingress_elapsed_ms=admitted.ingress_elapsed_ms,
            identification_elapsed_ms=admitted.identification_elapsed_ms,
            operation_elapsed_ms=(
                admitted.ingress_elapsed_ms
                + admitted.identification_elapsed_ms
                + (0 if cache_status == "HIT" else worker_result.elapsed_ms)
            ),
            deadline_stage=("PARSER" if worker_result.status == "TIMEOUT" else None),
        )
        if worker_result.status != "COMPLETE":
            return (
                self._failure(
                    worker_result.status,
                    admitted.scope_id,
                    admitted.relative_path,
                    admitted.source_sha256,
                    source_format,
                    backend_name,
                    resources,
                    worker_result.failure_reason_code
                    or (
                        "RUNTIME_DEPENDENCY_UNAVAILABLE"
                        if worker_result.status == "UNAVAILABLE"
                        else None
                    ),
                    failure_reason_code=worker_result.failure_reason_code,
                    failure_exception_type=worker_result.failure_exception_type,
                ),
                cache_status,
            )
        return (
            self._complete(
                admitted,
                worker_result,
                resources,
                source_format,
                backend_name,
            ),
            cache_status,
        )

    def _worker_for(
        self,
        source_format: str,
        profile: str,
        *,
        page_range: tuple[int, int] | None = None,
        query: str | None = None,
        native_view: str | None = None,
    ) -> tuple[str, IsolatedParserWorker] | None:
        if profile == "STRUCTURE_NATIVE" and source_format == "PDF":
            return (
                "PyMuPDFNativeStructure",
                IsolatedParserWorker(
                    worker_target=_pymupdf_pdf_structure_worker,
                    timeout_seconds=MAX_PARSER_ELAPSED_SECONDS,
                    memory_bytes=MAX_PDF_PARSER_MEMORY_BYTES,
                ),
            )
        if (
            profile == "MAP"
            and query is not None
            and source_format
            in {
                "PDF",
                "EPUB",
                "DOCX",
                "XLSX",
                "PPTX",
            }
        ):
            return (
                "STEWARDStreamingMap",
                IsolatedParserWorker(
                    worker_target=StreamingQueryMapWorker(source_format, query),
                    timeout_seconds=MAX_DEEP_PARSER_ELAPSED_SECONDS,
                    memory_bytes=MAX_PDF_PARSER_MEMORY_BYTES,
                ),
            )
        if profile == "FORMULA_NATIVE" and source_format == "XLSX":
            return (
                "openpyxl-formula-stream",
                IsolatedParserWorker(
                    worker_target=_openpyxl_xlsx_formula_worker,
                    timeout_seconds=MAX_DEEP_PARSER_ELAPSED_SECONDS,
                    memory_bytes=MAX_DEEP_PARSER_MEMORY_BYTES,
                ),
            )
        if profile == "OCR_NATIVE" and source_format == "PDF":
            return (
                "STEWARDPageOCR",
                IsolatedParserWorker(
                    worker_target=_rapidocr_pdf_worker,
                    timeout_seconds=MAX_DEEP_PARSER_ELAPSED_SECONDS,
                    memory_bytes=MAX_PDF_PARSER_MEMORY_BYTES,
                ),
            )
        if profile == "DEEP" and source_format == "PDF" and page_range is not None:
            return (
                "Docling",
                IsolatedDoclingWorker(worker_target=DoclingPageRangeWorker(*page_range)),
            )
        if profile == "OCR" and source_format == "PDF" and page_range is not None:
            return (
                "Docling",
                IsolatedDoclingWorker(worker_target=DoclingOcrPageRangeWorker(*page_range)),
            )
        if profile == "OCR" and source_format in {"PDF", *IMAGE_SOURCE_FORMATS}:
            return "Docling", self.macos_ocr_worker
        if profile == "ENRICHED" and source_format in {"PDF", *IMAGE_SOURCE_FORMATS}:
            return "Docling", self.enriched_worker
        if profile in {"DEEP", "ENRICHED"}:
            return "Docling", self.docling_worker
        if profile != "FAST":
            return None
        return {
            "PDF": ("PyMuPDF4LLM", self.worker),
            "EPUB": (
                "STEWARDNativeEpub",
                IsolatedParserWorker(
                    worker_target=NativeEpubWorker(native_view or "READ"),
                    timeout_seconds=MAX_DEEP_PARSER_ELAPSED_SECONDS,
                    memory_bytes=MAX_PDF_PARSER_MEMORY_BYTES,
                ),
            ),
            "DOCX": ("MarkItDown", self.docx_worker),
            "XLSX": ("openpyxl", self.xlsx_worker),
            "PPTX": ("python-pptx", self.pptx_worker),
        }.get(source_format)

    @staticmethod
    def _items_contain_query(items: tuple[NormalizedDocumentItem, ...], query: str) -> bool:
        return any(match_document_text(item.text_or_value, query) for item in items)

    @staticmethod
    def _items_contain_native_fidelity_query(
        items: tuple[NormalizedDocumentItem, ...], query: str
    ) -> bool:
        native_kinds = {
            "pdf_metadata",
            "pdf_outline",
            "pdf_annotation",
            "pdf_form_field",
            "docx_comment",
            "docx_footnote",
            "docx_endnote",
            "docx_revision",
            "xlsx_comment",
            "xlsx_chart",
            "pptx_speaker_notes",
            "pptx_accessibility",
            "pptx_chart",
        }
        return any(
            item.kind in native_kinds and match_document_text(item.text_or_value, query)
            for item in items
        )

    @staticmethod
    def _native_container(item: NormalizedDocumentItem, source_format: str) -> tuple[str, str, str]:
        location = item.location
        if source_format in {"PDF", *IMAGE_SOURCE_FORMATS}:
            page = location.get("page", 1 if source_format in IMAGE_SOURCE_FORMATS else None)
            if isinstance(page, int) and not isinstance(page, bool):
                return f"page:{page}", "PAGE", f"page {page}"
        if source_format == "XLSX":
            sheet = location.get("sheet")
            if isinstance(sheet, str):
                return f"sheet:{sheet}", "SHEET", f'sheet "{sheet}"'
        if source_format == "PPTX":
            slide = location.get("slide")
            if isinstance(slide, int) and not isinstance(slide, bool):
                return f"slide:{slide}", "SLIDE", f"slide {slide}"
        if source_format in {"DOCX", "EPUB"}:
            section = location.get("section")
            title = location.get("section_title")
            if isinstance(section, int) and not isinstance(section, bool):
                label = f"section {section}"
                if isinstance(title, str) and title:
                    label = f'{label} "{title}"'
                return f"section:{section}", "SECTION", label
        ordinal = location.get("ordinal")
        label = f"item {ordinal}" if isinstance(ordinal, int) else item.kind
        return "document", "DOCUMENT", label

    @staticmethod
    def _targeted_evidence_selection(
        items: tuple[NormalizedDocumentItem, ...],
        query: str,
        requested_page: int | None = None,
        *,
        source_format: str = "PDF",
        map_profile: str = "FAST",
    ) -> DocumentExecutionSelection:
        matched_indexes: list[int] = []
        matched_pages: set[int] = set()
        match_modes: set[str] = set()
        matched_containers: dict[str, tuple[str, str]] = {}
        container_items: dict[str, list[NormalizedDocumentItem]] = {}
        container_match_counts: dict[str, int] = {}
        for index, item in enumerate(items):
            container_id, container_kind, native_label = (
                StructuredDocumentParserAdapter._native_container(item, source_format)
            )
            container_items.setdefault(container_id, []).append(item)
            match = match_document_text(item.text_or_value, query)
            if match is None:
                continue
            match_modes.add(match.mode)
            matched_indexes.append(index)
            matched_containers[container_id] = (container_kind, native_label)
            container_match_counts[container_id] = container_match_counts.get(container_id, 0) + 1
            page = item.location.get("page")
            if isinstance(page, int) and not isinstance(page, bool) and page >= 1:
                matched_pages.add(page)
        pages = tuple(sorted(matched_pages))
        page_start: int | None = None
        page_end: int | None = None
        omitted: tuple[int, ...] = ()
        if pages:
            if requested_page is not None and requested_page not in pages:
                raise RuntimeFailure(
                    "INVALID_ARGUMENT", "requested evidence page does not match the query"
                )
            remaining_pages = tuple(
                page for page in pages if requested_page is None or page >= requested_page
            )
            if remaining_pages[-1] - remaining_pages[0] + 1 <= MAX_TARGETED_EVIDENCE_PAGES:
                page_start, page_end = remaining_pages[0], remaining_pages[-1]
            else:
                page_start = remaining_pages[0]
                page_end = page_start + MAX_TARGETED_EVIDENCE_PAGES - 1
                omitted = tuple(page for page in remaining_pages if page > page_end)
        elif requested_page is not None:
            raise RuntimeFailure(
                "INVALID_ARGUMENT", "requested evidence page does not match the query"
            )
        matched_container_ids = tuple(matched_containers)
        selected_container_ids = matched_container_ids
        omitted_container_ids: tuple[str, ...] = ()
        if source_format == "PDF" and page_start is not None:
            selected_container_ids = tuple(
                container_id
                for container_id in matched_container_ids
                if container_id.startswith("page:")
                and page_start <= int(container_id.split(":", 1)[1]) <= (page_end or page_start)
            )
            omitted_container_ids = tuple(
                container_id
                for container_id in matched_container_ids
                if container_id not in selected_container_ids
            )
        qualities = tuple(
            DocumentContainerQuality(
                container_id,
                matched_containers[container_id][0],
                matched_containers[container_id][1],
                container_match_counts[container_id],
                assess_document_quality(
                    container_items[container_id],
                    status="COMPLETE",
                    source_format=source_format,
                    view="READ",
                ),
            )
            for container_id in matched_container_ids
        )
        prefix = "STREAMING" if map_profile == "MAP" else "FAST"
        strategy = (
            "STREAMING_QUERY_MAP_PAGE_LOCAL_PROJECTION"
            if source_format == "PDF" and map_profile == "MAP"
            else (
                f"{prefix}_QUERY_MAP_THEN_TARGETED_DEEP_PAGE_RANGE"
                if source_format == "PDF"
                else f"{prefix}_QUERY_MAP_THEN_NATIVE_CONTAINER_QUALITY"
            )
        )
        return DocumentExecutionSelection(
            strategy,
            document_match_mode(match_modes),
            map_profile,
            len(items),
            len(matched_indexes),
            pages,
            page_start,
            page_end,
            omitted,
            matched_container_ids,
            selected_container_ids,
            omitted_container_ids,
            qualities,
        )

    @staticmethod
    def _attempt(
        profile: str,
        cache_status: str,
        observation: NormalizedDocumentObservation,
        source_format: str,
        view: str,
    ) -> DocumentExecutionAttempt:
        return DocumentExecutionAttempt(
            profile,
            observation.backend_name,
            observation.status,
            cache_status,
            assess_document_quality(
                observation.items,
                status=observation.status,
                source_format=source_format,
                view=view,
            ),
            observation.failure_reason_code,
            observation.failure_exception_type,
        )

    @staticmethod
    def _add_warning(
        observation: NormalizedDocumentObservation, warning: str
    ) -> NormalizedDocumentObservation:
        return replace(observation, warnings=tuple(dict.fromkeys((*observation.warnings, warning))))

    def _complete(
        self,
        admitted: _AdmittedDocumentSource,
        worker_result: _WorkerExecution,
        resources: DocumentResourceUsage,
        source_format: str,
        expected_backend_name: str,
        extra_warnings: tuple[str, ...] = (),
    ) -> NormalizedDocumentObservation:
        payload = worker_result.payload
        if payload is None:
            return self._failure(
                "PARSER_FAILED",
                admitted.scope_id,
                admitted.relative_path,
                admitted.source_sha256,
                source_format,
                expected_backend_name,
                resources,
            )
        items = _normalized_items(payload.get("items"))
        warnings_raw = payload.get("warnings", [])
        backend_name = payload.get("backend_name")
        backend_version = payload.get("backend_version")
        resource_extension = payload.get("resource_extension")
        continuation = payload.get("continuation")
        if (
            items is None
            or not isinstance(warnings_raw, list)
            or not all(isinstance(warning, str) for warning in warnings_raw)
            or not isinstance(backend_name, str)
            or not isinstance(backend_version, str)
            or backend_name != expected_backend_name
            or (resource_extension is not None and not isinstance(resource_extension, dict))
            or (continuation is not None and not isinstance(continuation, dict))
        ):
            return self._failure(
                "PARSER_FAILED",
                admitted.scope_id,
                admitted.relative_path,
                admitted.source_sha256,
                source_format,
                expected_backend_name,
                resources,
            )
        items = _format_native_items(items, source_format)
        warnings = [*warnings_raw, *extra_warnings]
        candidate_size = len(
            canonical_json(
                {
                    "source_format": source_format,
                    "backend_name": backend_name,
                    "backend_version": backend_version,
                    "warnings": warnings,
                    "items": [item.payload() for item in items],
                }
            )
        )
        final_resources = DocumentResourceUsage(
            resources.source_bytes,
            resources.expanded_bytes,
            resources.parser_elapsed_ms,
            resources.parser_memory_bytes,
            len(items),
            candidate_size,
            resources.admission_profile,
            resources.source_limit_bytes,
            resources.expanded_limit_bytes,
            resources.archive_member_count,
            resources.parser_timeout_limit_ms,
            resources.parser_memory_limit_bytes,
            resources.ingress_elapsed_ms,
            resources.identification_elapsed_ms,
            resources.operation_elapsed_ms,
            resources.deadline_stage,
            resource_extension if isinstance(resource_extension, dict) else resources.media,
        )
        if len(items) > MAX_PARSED_ITEMS_OR_BLOCKS or candidate_size > MAX_NORMALIZED_OUTPUT_BYTES:
            return self._failure(
                "RESOURCE_LIMIT",
                admitted.scope_id,
                admitted.relative_path,
                admitted.source_sha256,
                source_format,
                backend_name,
                final_resources,
            )
        return NormalizedDocumentObservation(
            "COMPLETE",
            source_format,
            backend_name,
            backend_version,
            DocumentSourceProvenance(
                admitted.scope_id,
                admitted.relative_path,
                admitted.source_sha256,
                (
                    CURRENT_FILESYSTEM_AUDIO
                    if source_format in AUDIO_SOURCE_FORMATS
                    else (
                        CURRENT_FILESYSTEM_VIDEO
                        if source_format in VIDEO_SOURCE_FORMATS
                        else CURRENT_FILESYSTEM_DOCUMENT
                    )
                ),
            ),
            tuple(warnings),
            items,
            final_resources,
            continuation=continuation if isinstance(continuation, dict) else None,
        )

    @staticmethod
    def _failure(
        status: str,
        scope_id: str,
        relative_path: str,
        source_sha256: str | None,
        source_format: str | None,
        backend_name: str | None,
        resources: DocumentResourceUsage,
        identification_reason: str | None = None,
        *,
        failure_reason_code: str | None = None,
        failure_exception_type: str | None = None,
    ) -> NormalizedDocumentObservation:
        return NormalizedDocumentObservation(
            status,
            source_format,
            backend_name,
            None,
            DocumentSourceProvenance(scope_id, relative_path, source_sha256),
            (),
            (),
            resources,
            identification_reason,
            failure_reason_code=failure_reason_code,
            failure_exception_type=failure_exception_type,
        )
