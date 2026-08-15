"""Deterministic offline coverage for the Structured Document XLSX slice."""

from __future__ import annotations

from dataclasses import dataclass
import os
from pathlib import Path
from time import sleep
from typing import Any
import zipfile

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo
import pytest

import local_steward.file_agent.runtime.structured_documents as structured_documents
from local_steward.file_agent.runtime import (
    CURRENT_FILESYSTEM_DOCUMENT,
    MAX_NORMALIZED_OUTPUT_BYTES,
    MAX_PARSED_ITEMS_OR_BLOCKS,
    IsolatedXlsxWorker,
    ProjectOwnedBoundedDocumentIngress,
    ScopeBinding,
    ScopeBindings,
    StructuredDocumentParserAdapter,
    identify_document_format,
)
from local_steward.file_agent.runtime.runtime import RuntimeFailure


def _bindings(tmp_path: Path) -> tuple[Path, ScopeBindings]:
    root = tmp_path / "isolated-workbooks"
    root.mkdir(parents=True)
    return root, ScopeBindings((ScopeBinding("managed", root),), (str(root),), ("managed",))


def _arguments(path: str = "sample.xlsx") -> dict[str, object]:
    return {"scope_id": "managed", "relative_path": path}


def _write_xlsx(path: Path, *, cells: int = 3, large_value: str | None = None) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.append(["Label", "Amount", "Formula"])
    sheet.append(["January", 10, "=B2*2"])
    sheet.append(["February", 20, "=SUM(B2:B3)"])
    sheet["A1"].comment = Comment("reviewed workbook header", "Analyst")
    if large_value is not None:
        for index, offset in enumerate(range(0, len(large_value), 30_000), start=2):
            sheet.cell(row=index, column=4, value=large_value[offset : offset + 30_000])
    elif cells > 3:
        for index in range(cells):
            sheet.cell(row=(index // 64) + 5, column=(index % 64) + 1, value=f"V{index}")
    table = Table(displayName="SummaryTable", ref="A1:C3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    sheet.add_chart(chart, "E2")
    detail = workbook.create_sheet("Detail")
    detail["A1"] = "detail marker"
    workbook.save(path)
    workbook.close()


def _worker_payload(items: list[dict[str, object]]) -> dict[str, Any]:
    return {
        "backend_name": "openpyxl",
        "backend_version": "3.1.5",
        "warnings": [],
        "items": items,
    }


def _sleep_worker(_path: str) -> dict[str, Any]:
    sleep(2.0)
    return _worker_payload([])


def _crash_worker(_path: str) -> dict[str, Any]:
    os._exit(19)


@dataclass
class _NeverWorker:
    def run(self, _source_path: Path):
        raise AssertionError("rejected XLSX input must not reach an adapter worker")


def _adapter(
    tmp_path: Path, xlsx_worker: object | None = None
) -> tuple[Path, StructuredDocumentParserAdapter]:
    root, bindings = _bindings(tmp_path)
    adapter = StructuredDocumentParserAdapter(ProjectOwnedBoundedDocumentIngress(bindings))
    if xlsx_worker is not None:
        adapter.xlsx_worker = xlsx_worker  # type: ignore[assignment]
    return root, adapter


def _malformed_xlsx(path: Path) -> None:
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("[Content_Types].xml", "<Types />")
        archive.writestr("xl/workbook.xml", "<workbook>")


def _container(path: Path, members: dict[str, bytes]) -> None:
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, value in members.items():
            archive.writestr(name, value)


def test_real_openpyxl_worker_preserves_workbook_sheet_cell_formula_table_and_chart_semantics(
    tmp_path: Path,
) -> None:
    root, adapter = _adapter(tmp_path)
    _write_xlsx(root / "sample.xlsx")

    observation = adapter.observe(_arguments())
    payload = observation.payload()
    items = [item.payload() for item in observation.items]

    assert observation.status == "COMPLETE"
    assert observation.source_format == "XLSX"
    assert observation.backend_name == "openpyxl"
    assert observation.backend_version == "3.1.5"
    assert observation.provenance.payload()["source_kind"] == CURRENT_FILESYSTEM_DOCUMENT
    assert observation.provenance.relative_path == "sample.xlsx"
    assert observation.resources.expanded_bytes > 0
    assert {item["kind"] for item in items} >= {
        "xlsx_workbook",
        "xlsx_sheet",
        "xlsx_cell",
        "xlsx_table",
        "xlsx_chart",
        "xlsx_comment",
    }
    formula = next(item for item in items if item.get("location", {}).get("cell") == "C2")
    assert formula["text_or_value"] == "=B2*2"
    assert formula["extension"]["formula"] == "=B2*2"
    assert formula["role"] == "FORMULA"
    assert formula["extension"]["row_hidden"] is False
    assert formula["extension"]["column_hidden"] is False
    comment = next(item for item in items if item["kind"] == "xlsx_comment")
    assert comment["text_or_value"] == "reviewed workbook header"
    assert comment["extension"]["author"] == "Analyst"
    chart = next(item for item in items if item["kind"] == "xlsx_chart")
    assert chart["extension"]["evaluated"] is False
    assert chart["extension"]["series_references"] == [
        {
            "name_reference": "'Summary'!B1",
            "value_reference": "'Summary'!$B$2:$B$3",
        }
    ]
    assert {item["location"].get("sheet") for item in items if "sheet" in item["location"]} >= {
        "Summary",
        "Detail",
    }
    assert all(isinstance(item, dict) for item in payload["items"])  # type: ignore[arg-type]
    assert "Workbook(" not in str(payload)


def test_valid_xlsx_signature_routes_despite_suffix_and_unknown_zip_is_rejected_before_worker(
    tmp_path: Path,
) -> None:
    root, adapter = _adapter(tmp_path)
    _write_xlsx(root / "misleading.data")
    rejected_root, rejected_adapter = _adapter(tmp_path / "rejected", _NeverWorker())
    _container(rejected_root / "random.xlsx", {"ordinary.txt": b"not an OOXML workbook"})

    valid = adapter.observe(_arguments("misleading.data"))
    rejected = rejected_adapter.observe(_arguments("random.xlsx"))

    assert valid.status == "COMPLETE" and valid.source_format == "XLSX"
    assert rejected.status == "UNSUPPORTED_FORMAT"
    assert rejected.identification_reason == "UNACCEPTED_FORMAT"
    assert rejected.items == ()
    assert identify_document_format(b"random", "false.xlsx").reason == "FORMAT_MISMATCH"


def test_recognized_malformed_xlsx_is_never_published(tmp_path: Path) -> None:
    root, adapter = _adapter(tmp_path)
    _malformed_xlsx(root / "broken.xlsx")

    observation = adapter.observe(_arguments("broken.xlsx"))

    assert observation.status == "MALFORMED"
    assert observation.items == () and observation.warnings == ()


def test_source_and_expanded_container_limits_reject_before_xlsx_worker_dispatch(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    root, adapter = _adapter(tmp_path, _NeverWorker())
    configured_limit = 32
    adapter.ingress.max_staged_bytes = configured_limit
    monkeypatch.setattr(structured_documents, "MAX_XLSX_EXPANDED_BYTES", configured_limit)
    (root / "large.xlsx").write_bytes(b"x" * (configured_limit + 1))
    _container(
        root / "expanded.xlsx",
        {
            "[Content_Types].xml": b"<Types />",
            "xl/workbook.xml": b"<workbook />",
            "xl/oversized.xml": b"x" * (configured_limit + 1),
        },
    )

    source_limited = adapter.observe(_arguments("large.xlsx"))
    adapter.ingress.max_staged_bytes = 4_096
    expanded_limited = adapter.observe(_arguments("expanded.xlsx"))

    assert source_limited.status == "RESOURCE_LIMIT" and source_limited.items == ()
    assert expanded_limited.status == "RESOURCE_LIMIT" and expanded_limited.items == ()
    assert expanded_limited.resources.expanded_bytes > configured_limit


def test_unsafe_container_member_path_is_rejected_without_publication(tmp_path: Path) -> None:
    root, adapter = _adapter(tmp_path, _NeverWorker())
    _container(
        root / "unsafe.xlsx",
        {
            "[Content_Types].xml": b"<Types />",
            "xl/workbook.xml": b"<workbook />",
            "../outside.xml": b"escape",
        },
    )

    observation = adapter.observe(_arguments("unsafe.xlsx"))

    assert observation.status == "RESOURCE_LIMIT"
    assert observation.identification_reason == "UNSAFE_CONTAINER_PATH"
    assert observation.items == ()


@pytest.mark.parametrize(
    ("fixture", "expectation"),
    (
        ("many", "items"),
        ("large-output", "bytes"),
    ),
)
def test_real_xlsx_normalized_item_and_output_limits_publish_no_partial_workbook(
    tmp_path: Path, fixture: str, expectation: str
) -> None:
    root, adapter = _adapter(tmp_path)
    if fixture == "many":
        _write_xlsx(root / "many.xlsx", cells=MAX_PARSED_ITEMS_OR_BLOCKS)
    else:
        _write_xlsx(root / "large-output.xlsx", large_value="x" * (MAX_NORMALIZED_OUTPUT_BYTES + 1))

    observation = adapter.observe(_arguments(f"{fixture}.xlsx"))

    assert observation.status == "RESOURCE_LIMIT"
    assert observation.items == ()
    if expectation == "items":
        assert observation.resources.parsed_items_or_blocks > MAX_PARSED_ITEMS_OR_BLOCKS
    else:
        assert observation.resources.normalized_output_bytes > MAX_NORMALIZED_OUTPUT_BYTES


@pytest.mark.parametrize(
    ("target", "timeout", "memory", "expected"),
    (
        (_sleep_worker, 0.25, 640 * 1024 * 1024, "TIMEOUT"),
        (_sleep_worker, 2.0, 1, "RESOURCE_LIMIT"),
        (_crash_worker, 2.0, 640 * 1024 * 1024, "PARSER_FAILED"),
    ),
)
def test_xlsx_worker_uses_existing_isolation_timeout_and_failure_mapping(
    tmp_path: Path, target: object, timeout: float, memory: int, expected: str
) -> None:
    root, _bindings_value = _bindings(tmp_path)
    _write_xlsx(root / "sample.xlsx")

    result = IsolatedXlsxWorker(
        worker_target=target, timeout_seconds=timeout, memory_bytes=memory
    ).run(  # type: ignore[arg-type]
        root / "sample.xlsx"
    )

    assert result.status == expected


def test_parser_timeout_publishes_stage_and_measured_operation_budget(tmp_path: Path) -> None:
    worker = IsolatedXlsxWorker(worker_target=_sleep_worker, timeout_seconds=0.1)
    root, adapter = _adapter(tmp_path, worker)
    _write_xlsx(root / "sample.xlsx")

    observation = adapter.observe(_arguments())

    assert observation.status == "TIMEOUT"
    assert observation.items == ()
    assert observation.resources.deadline_stage == "PARSER"
    assert observation.resources.parser_timeout_limit_ms == 100
    assert observation.resources.parser_elapsed_ms >= 100
    assert observation.resources.operation_elapsed_ms >= observation.resources.parser_elapsed_ms
    assert observation.resources.ingress_elapsed_ms >= 0
    assert observation.resources.identification_elapsed_ms >= 0


def test_external_relationships_are_observed_as_safe_ignored_metadata_and_not_followed(
    tmp_path: Path,
) -> None:
    root, adapter = _adapter(tmp_path)
    path = root / "external.xlsx"
    _write_xlsx(path)
    with zipfile.ZipFile(path, "a") as archive:
        archive.writestr(
            "custom.rels",
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" TargetMode="External" Target="https://example.invalid/workbook" />'
            "</Relationships>",
        )

    observation = adapter.observe(_arguments("external.xlsx"))

    assert observation.status == "COMPLETE"
    assert observation.warnings == ("external_relationships_ignored:1",)


def test_scope_binding_rejects_escape_and_pdf_identification_regression_is_unchanged(
    tmp_path: Path,
) -> None:
    root, adapter = _adapter(tmp_path, _NeverWorker())
    outside = tmp_path / "outside.xlsx"
    _write_xlsx(outside)
    (root / "escape.xlsx").symlink_to(outside)

    with pytest.raises(RuntimeFailure, match="SCOPE_BINDING_FAILED"):
        adapter.observe(_arguments("../outside.xlsx"))
    assert adapter.observe(_arguments("escape.xlsx")).status == "UNAVAILABLE"
    assert identify_document_format(b"%PDF-1.7\n", "still.xlsx").source_format == "PDF"


def test_xlsx_worker_explicitly_keeps_links_and_macros_disabled_and_formulas_unexecuted() -> None:
    source = (
        Path(__file__).parents[1] / "src/local_steward/file_agent/runtime/structured_documents.py"
    )
    text = source.read_text(encoding="utf-8")
    assert "data_only=False" in text
    assert "keep_links=False" in text
    assert "keep_vba=False" in text


def test_auto_formula_view_uses_native_stream_and_preserves_structured_references(
    tmp_path: Path,
) -> None:
    root, adapter = _adapter(tmp_path)
    path = root / "structured.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Calculations"
    sheet.append(["Amount", "Tax"])
    sheet.append([10, "=[@Amount]*0.2"])
    table = Table(displayName="SalesTable", ref="A1:B2")
    sheet.add_table(table)
    sheet["D2"] = "=SUM(SalesTable[Amount])"
    workbook.save(path)
    workbook.close()

    observation = adapter.observe(
        _arguments("structured.xlsx")
        | {"parser_profile": "AUTO", "view": "FORMULAS", "intent": "FORMULAS"}
    )

    assert observation.status == "COMPLETE"
    assert observation.backend_name == "openpyxl-formula-stream"
    assert observation.execution is not None
    assert observation.execution.initial_profile == "FORMULA_NATIVE"
    assert {item.location["cell"] for item in observation.items} == {"B2", "D2"}
    structured = next(item for item in observation.items if item.location["cell"] == "D2")
    assert structured.role == "FORMULA"
    assert structured.extension is not None
    assert structured.extension["formula_kind"] == "STRUCTURED_REFERENCE"
    assert structured.extension["structured_references"] == ["SalesTable[Amount]"]
    assert structured.extension["evaluated"] is False
    assert observation.resources.deadline_stage is None
    assert observation.resources.operation_elapsed_ms >= observation.resources.parser_elapsed_ms
